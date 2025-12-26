import multiprocessing
import os
import sys
import copy
import tkinter

import numpy as np
import subprocess

from time import time, sleep
from datetime import datetime
from pathlib import Path
from multiprocessing import Pool, cpu_count
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from enum import Enum, auto

from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from tkinter.ttk import Treeview, Progressbar, Style
from ttkwidgets import CheckboxTreeview

import json
import git
import shutil

################################################################################################################
# ENUM
# --------------------------------------------------------------------------------------------------------------
# ENUM : Preset save popup state
class PresetSavePopupState(Enum):
    NORMAL = auto()
    SAVE = auto()
    EXIT = auto()
    NO_NAME = auto()
    NO_FILE = auto()


# --------------------------------------------------------------------------------------------------------------
# ENUM : Preset load popup state
class PresetLoadPopupState(Enum):
    NORMAL = auto()
    LOAD = auto()
    EXIT = auto()
    NO_NAME = auto()
    DELETE = auto()


# --------------------------------------------------------------------------------------------------------------
# ENUM : Preset load popup state
class EditPopupState(Enum):
    NORMAL = auto()
    SAVE = auto()
    EXIT = auto()
    NO_FILE = auto()


# --------------------------------------------------------------------------------------------------------------
# ENUM : Preset load popup state
class MainUIState(Enum):
    NORMAL = auto()
    DISABLED = auto()

################################################################################################################
# OS RELATED
# required packages : numpy, openyxl, tkinter, ttkwidgets, gitpython, pywinauto(MS win only)
# - macOS에서는 path 구분자가 "\"가 아닌 "/"임
_SEP = os.sep

################################################################################################################
# SUB FUNCTION
# --------------------------------------------------------------------------------------------------------------
# SUB FUNCTION : String to list by key slicing
def ParseToList(sParam: str, sKey: str, bUpper = False, bTrim = False, bCustom = False):
    res = list()
    tList = list()
    
    if bCustom:
        tList = CustomSplit(sParam, sKey)
    else:
        tList = sParam.split(sKey)
    
    for tKey in tList:
        tStr = tKey
        if bUpper: tStr = tStr.upper()
        if bTrim: tStr = tStr.strip()
        res.append(tStr)
    
    return res


# --------------------------------------------------------------------------------------------------------------
# SUB FUNCTION : Custom Split with <>
def CustomSplit(sParam: str, sKey: str):
    bValue = False
    cStr = ""
    tStr = ""
    res = list()
    
    for x in range(0, len(sParam)):
        cStr = sParam[x:x+1]
        if cStr == "<":
            bValue = True
        elif cStr == ">":
            bValue = False
        
        if cStr == sKey and bValue == False:
            res.append(tStr)
            tStr = ""
        else:
            tStr = tStr + cStr
    
    res.append(tStr)
    
    return res


# --------------------------------------------------------------------------------------------------------------
# SUB FUNCTION : Get Excel Address with x & y
def Xlref(row, column, zero_indexed=True):
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)


# --------------------------------------------------------------------------------------------------------------
# SUB FUNCTION : Data Type Check
def TypeCheck(sType: str, sValue: str):
    if len(sValue) > 0:
        match sType:
            case "int" | "long":
                try:
                    if(str(int(sValue))) != sValue: return "INT_VALUE_ERROR"
                except ValueError:
                    return "INT_VALUE_ERROR"
            case "float" | "double" :
                try:                    
                    if(str(float(sValue))) != sValue:
                        if(str(int(sValue))) != sValue:
                            return "FLOAT_VALUE_ERROR"
                except ValueError:
                    return "FLOAT_VALUE_ERROR"                                    
            case "bool":
                if sValue.lower() != "false" and sValue.lower() != "true": return "BOOL_VALUE_ERROR"
            case "type":
                if " " in sValue: return "TYPE_VALUE_SPCAE_ERROR"
            case "string":
                if "\"" in sValue: return "STRING_VALUE_QUOTE_ERROR"                
    else:
        match sType:
            case "int" | "long" : return "INT_VALUE_NULL"
            case "float" | "double" : return "FLOAT_VALUE_NULL"
            case "bool": return "BOOL_VALUE_NULL"
            case "type": return "TYPE_VALUE_NULL"
    
    return "PASS"


# --------------------------------------------------------------------------------------------------------------
# SUB FUNCTION : List to String
def ListToString(l: list, split: str):
    bFirst = True
    res = ""
    for tKey in l:
        if bFirst is True:
            res = str(tKey)
            bFirst = False
        else:
            res = res + split + str(tKey)
        
    return res


# --------------------------------------------------------------------------------------------------------------
# SUB FUNCTION : Get tab string for match align
def GetTabStr(baseLen: int, usedLen: int, offset: int):
    res = ""
    
    if baseLen > usedLen:
        for i in range(1, int(baseLen / 4) - int((usedLen + offset) / 4)):
            res = res + "\t"
    else:
        res = "\t"
    
    return res


################################################################################################################
# EXPORT JSON CLASS & FUNCTION
# --------------------------------------------------------------------------------------------------------------
# JSON CLASS : Array size manager
class cArraySizeRefCheck():
    def __init__(self):
        self.dArraySize = dict()
    
    def AddArraySize(self, sheetName: str, y: int, x: int, arrSize: int):
        referName = sheetName + "/" + str(x)
        if referName in self.dArraySize:
            self.dArraySize[referName][y] = arrSize
        else:
            dArraySizeMember = dict()
            dArraySizeMember[y] = arrSize
            self.dArraySize[referName] = dArraySizeMember
    
    def GetSize(self, sheetName: str, y: int, x: int):
        referName = sheetName + "/" + str(x)
        if referName in self.dArraySize:
            if y in self.dArraySize[referName]:
                return self.dArraySize[referName][y]
        
        return -1


# --------------------------------------------------------------------------------------------------------------
# JSON CLASS : Enum type list manage
class cEnumList:
    def __init__(self):
        self.dEnum = dict()
        self.dEnumList = dict()
        self.dEnumNocode = dict()
        self.dComment = dict()
        self.dTextCheck = dict()
        self.dValueCheck = dict()
    
    def AddType(self, sheetName: str, sEnum: str, sText: str, sValue: int, sComment: str, bNocode : bool):
        sEnum = sEnum.upper()
        sText = sText.upper()
        sValue = sValue.upper()
        
        # Null Check
        if len(sEnum) == 0: return "NULL_TYPE_GROUP"
        if len(sText) == 0: return "NULL_TYPE_TEXT"
        if len(sValue) == 0: return "NULL_TYPE_VALUE"

        if sEnum in self.dEnumList:
            if self.dEnumList[sEnum] != sheetName: return "DUPLICATE_TYPE_GROUP_ANOTHER_SHEET"
        else:
            self.dEnumList[sEnum] = sheetName

        self.dEnumNocode[sEnum] = bNocode
        dValue = dict()
        if sEnum in self.dEnum:
            dValue = self.dEnum[sEnum]
        
        if sText in dValue: return "DUPLICATE_TYPE_TEXT_IN_TYPE_GROUP"
        
        if sEnum + "/" + sValue in self.dValueCheck: return "DUPLICATE_TYPE_VALUE_IN_TYPE_GROUP"
        
        dValue[sText] = sValue
        self.dEnum[sEnum] = dValue
        self.dComment[sEnum + "/" + sValue] = sComment
        self.dTextCheck[sEnum + "/" + sText] = sValue
        self.dValueCheck[sEnum + "/" + sValue] = sText
        
        return "PASS"


# --------------------------------------------------------------------------------------------------------------
# JSON CLASS : Log data
class cLog:
    def __init__(self):
        self.fullPath = ""
        self.fileName = ""        
        self.sheetName = ""
        self.colName = ""
        self.y = 0
        self.x = 0
        self.msg = ""


# --------------------------------------------------------------------------------------------------------------
# JSON CLASS : Log list manage
class cLogList:
    def __init__(self):
        self.logs = dict()
    
    def Add(self, fullPath: str, fileName: str, sheetName: str, colName: str, y: int, x: int, msg: str):
        log = cLog()
        log.fullPath = fullPath
        log.fileName = fileName
        log.sheetName = sheetName
        log.colName = colName
        log.y = y + 1
        log.x = x + 1
        log.msg = msg
        
        no = len(self.logs) + 1
        self.logs[no] = log

    def Merge(self, fullPath: str, fileName: str, sheetName: str, colName: str, y: int, x: int, msg: str):
        log = cLog()
        log.fullPath = fullPath
        log.fileName = fileName
        log.sheetName = sheetName
        log.colName = colName
        log.y = y
        log.x = x
        log.msg = msg
        
        no = len(self.logs) + 1
        self.logs[no] = log
    
    def LogCount(self):
        return len(self.logs)
    
    def LogPrint(self):
        for i in self.logs:
            log = self.logs[i]
            print(i, log.fileName, log.sheetName, log.colName, log.y, log.x, log.msg)


# --------------------------------------------------------------------------------------------------------------
# JSON CLASS : Table schema
class cTableSchema:
    def __init__(self):
        self.fileName = ""
        self.filePath = ""
        self.sheetName = ""
        self.subFolder = ""

        self.ws = ""
        
        self.hdTitle = dict()
        self.hdParam = dict()
        self.hdExport = dict()
        self.dCheckedX = dict()
        
        self.dArray = dict()
        self.dOrigin = dict()
        self.dUnique = dict()
        
        self.dDataType = dict()
        self.dEnum = dict()
        self.dEnumRef = dict()
        self.dValidCheck = dict()
        self.dChange = dict()
        self.dChanged = dict()
        self.dPrefix = dict()
        self.dSuffix = dict()
        self.dNull = dict()
        self.dArraySize = dict()
        self.dArraySizeRef = dict()
        self.dReplaceComma = dict()
        
        self.bNeedEnum = False
        self.bNeedText = False
        self.bNocode = False
        self.bValueOnly = False
        self.bNeedMerge = list()
        self.refSheet = list()
        
        self.dArraySizeOrg = dict()
        self.dArraySizeCnt = dict()
        
        self.cExportX = list()
        self.cCheckX = list()
        
        self.typeSheet = ""
        self.mergeData = list()
        self.posKey = 0
        self.maxRow = 0
        self.maxCol = 0
        
        self.logList = list()
       
    def Init(self, ws, fPath: str, fName: str, sName: str, dMergeList: dict):
        self.ws = ws        
        self.fileName = fName
        self.fullPath = fPath
        self.sheetName = sName
        
        self.maxRow = len(self.ws)
        self.maxCol = len(self.ws[0])
        
        self.ReadHeader()
        self.ParseHeader(dMergeList)
        self.HeaderErrorCheck()

    def ReadHeader(self):
        self.typeSheet = self.ws[1][0]
        self.mergeData = list()
        self.dataOption = list()

        if self.typeSheet.lower() == "#merge":
            if self.ws[2][0] is not None:
                self.mergeData = ParseToList(self.ws[2][0], ",")

        if self.typeSheet == "#data" or self.typeSheet == "#type":
            if self.ws[2][0] is not None:
                self.dataOption = ParseToList(self.ws[2][0], ",")
                for option in self.dataOption :
                    if option == "nocode" :
                        self.bNocode =  True
                    if option == "valueonly" :
                        self.bValueOnly = True
            if self.ws[0][0] is not None:
                self.subFolder = self.ws[0][0]

        dRemoveX = dict()
        
        for tKey in range(1, self.maxCol):
            if self.ws[1][tKey] is None:
                dRemoveX[tKey] = tKey
            else:
                self.hdTitle[tKey] = self.ws[1][tKey]
                self.dCheckedX[tKey] = tKey
        
        for tKey in range(1, self.maxCol):
            if self.ws[2][tKey] is None:
                dRemoveX[tKey] = tKey
                if tKey in self.dCheckedX: del(self.dCheckedX[tKey])
            elif tKey in self.dCheckedX:
                self.hdParam[tKey] = self.ws[2][tKey]

        for tKey in range(1, self.maxCol):
            if self.ws[3][tKey] is None:
                dRemoveX[tKey] = tKey
                if tKey in self.dCheckedX: del(self.dCheckedX[tKey])
            elif tKey in self.dCheckedX:
                self.hdExport[tKey] = self.ws[3][tKey]
        
        for tKey in dRemoveX:
            if tKey in self.hdTitle: del(self.hdTitle[tKey])
            if tKey in self.hdParam: del(self.hdParam[tKey])
            if tKey in self.hdExport: del(self.hdExport[tKey])
    
    def ParseHeader(self, dMergeList: dict):
        for tPos in self.dCheckedX:
            sTitle = self.hdTitle[tPos]
            if sTitle[len(sTitle)-2:len(sTitle)] == "[]":
                sTitle = sTitle[:len(sTitle)-2]
                self.hdTitle[tPos] = sTitle
                self.dArray[tPos] = sTitle
        
            self.hdParam[tPos] = ParseToList(self.hdParam[tPos], ",", False, True, True)
            
            for tStr in self.hdParam[tPos]:
                tStr = str(tStr).lower()
                
                match tStr:
                    case "key":
                        self.posKey = tPos
                        if tPos not in self.dOrigin: self.dOrigin[tPos] = str(self.sheetName + "/" + self.hdTitle[tPos]).lower()
                        if tPos not in self.dUnique: self.dUnique[tPos] = sTitle                        
                    case "int" | "float" | "string" | "bool" | "stream" | "long" | "double":
                        if tPos not in self.dDataType: self.dDataType[tPos] = tStr
                    case "localtext":
                        if tPos not in self.dDataType: self.dDataType[tPos] = tStr
                        self.bNeedText = True
                    case "origin":
                        if tPos not in self.dOrigin: self.dOrigin[tPos] = str(self.sheetName + "/" + self.hdTitle[tPos]).lower()
                    case "unique":
                        if tPos not in self.dUnique: self.dUnique[tPos] = sTitle
                    case "null":
                        if tPos not  in self.dNull: self.dNull[tPos] = "NULL_CHECK"
                
                nList = ParseToList(tStr, "<")                
                if len(nList) == 2:
                    if len(nList[0]) > 0 and len(nList[1]) > 1:
                        strKey = nList[0]
                        strParam = nList[1][0:len(nList[1])-1]
                        match strKey:
                            case "type":
                                self.bNeedEnum = True                                
                                self.dEnum[tPos] = strParam.upper()
                                self.dDataType[tPos] = "type"
                            case "typeref":                                
                                self.dEnum[tPos] = strParam
                                self.dEnumRef[tPos] = strParam
                                self.dDataType[tPos] = "string"
                            case "check":
                                self.dValidCheck[tPos] = strParam
                                nList = ParseToList(strParam, "/")
                                if len(nList) == 2:
                                    if nList[0] in dMergeList and nList[0] not in self.bNeedMerge:
                                        self.bNeedMerge.append(nList[0])
                                    elif nList[0] not in self.refSheet:
                                        self.refSheet.append(nList[0])
                                else:
                                    self.logList.append("INVALID_REF_NAME," + str(tPos))    # 참조 필드명 없음
                            case "change":
                                cPos = self.GetPosByName(strParam)
                                if cPos > 0:
                                    self.dChange[tPos] = cPos
                                    self.dChanged[tPos] = cPos
                                else:
                                    self.logList.append("INVALID_CHANGE_FILED," + str(tPos))    # 변환 필드명 없음
                            case "prefix":
                                self.dPrefix[tPos] = strParam
                            case "suffix":
                                self.dSuffix[tPos] = strParam
                            case "arraysize":
                                self.dArraySize[tPos] = strParam
                            case "arraysizeref":
                                cPos = self.GetPosByName(strParam)
                                if cPos > 0:
                                    self.dArraySizeRef[tPos] = cPos
                                    self.dArraySizeOrg[cPos] = strParam
                                else:
                                    self.logList.append("INVALID_ARRAYREF_NAME," + str(tPos))    # 배열 길이 참조 필드명 없음
                            case "arraysizecnt":
                                cPos = self.GetPosByName(strParam)
                                if cPos > 0:
                                    self.dArraySizeRef[tPos] = cPos
                                    self.dArraySizeCnt[cPos] = strParam
                                else:
                                    self.logList.append("INVALID_ARRAYCNTREF_NAME," + str(tPos))    # 배열 길이 참조 필드명 없음
                            case "replacecomma":
                                self.dReplaceComma[tPos] = strParam
            
            match self.hdExport[tPos].lower():
                case "data": self.cExportX.append(tPos)
                case "check": self.cCheckX.append(tPos)
    
    def HeaderErrorCheck(self):
        if len(self.cExportX) == 0:
            self.logList.append("NO_CHECKED,1") # 추출할 데이터가 아무것도 없음
        
        if self.posKey == 0 and not self.bValueOnly:
            self.logList.append("NO_KEY_FIELD,1") # 키 필드가 정의되어 있지 않음

        if self.subFolder != "" and self.subFolder[-1] != '/' :
            self.logList.append("NOT_SUB_FOLDER,1")  # 서브 폴더 마지막이 / 아님

        dTitle = dict()
        
        for tPos in self.dCheckedX:
            sTitle = self.hdTitle[tPos]
            
            if ' ' in sTitle:
                self.logList.append("SPACE_IN_TITLE," + str(tPos)) # 필드 이름에 공백이 포함되어 있음
            
            if sTitle in dTitle:
                self.logList.append("DUPLICATE_TITLE," + str(tPos)) # 필드 이름 중복 사용
            else:
                dTitle[sTitle] = tPos
            
            if tPos not in self.dDataType:
                self.logList.append("NO_DATA_TYPE," + str(tPos)) # 데이터 타입이 정의되어 있지 않음
            
            if tPos in self.dNull:
                if tPos in self.dEnum:
                    self.logList.append("ENUM_NULL_TO_SPACE," + str(tPos)) # ENUM에서는 NULL을 빈 값으로 변환하면 안 됨
                if tPos == self.posKey:
                    self.logList.append("EMPTY_KEY," + str(tPos)) # 키 필드는 빈 값이 있으면 안 됨
                if tPos in self.dOrigin:
                    self.logList.append("EMPTY_REFERECNE_VALUE," + str(tPos)) # 참조되는 값에는 빈 값으로 변환하면 안 됨
                if tPos in self.dUnique:
                    self.logList.append("EMPTY_UNIQUE_VALUE," + str(tPos)) # 고유 값에는 빈 값으로 변환하면 안 됨
        
        for tPos in self.dEnumRef:
            sTitle = self.hdTitle[tPos]
            x = self.GetPosByName(self.dEnumRef[tPos])
            
            if x > 0:
                self.dEnumRef[tPos] = x
            else:
                self.logList.append("UNVALID_REFRENCE_FIELD," + str(tPos)) # 존재하지 않는 필드를 참조하면 안 됨
        
    def GetPosByName(self, sTitle: str):
        for tPos in self.hdTitle:
            if self.hdTitle[tPos] == sTitle:
                return tPos
        
        return -1


# --------------------------------------------------------------------------------------------------------------
# JSON CLASS : Sheet data manager (y, x)
class cSheetData():
    def __init__(self):
        self.data = dict()
    
    def Add(self, y: int, x: int, value):
        if y in self.data:
            self.data[y][x] = value
        else:
            dRow = dict()
            dRow[x] = value
            self.data[y] = dRow
    
    def Find(self, y: int, x: int):
        if y in self.data:
            if x in self.data[y]:
                return self.data[y][x]
        
        return ""
    
    def Exists(self, y: int, x: int):
        if y in self.data:
            if x in self.data[y]:
                return True
        
        return False


# --------------------------------------------------------------------------------------------------------------
# JSON CLASS : Timer data
class cTimer():
    def __init__(self):
        self.tTotal = 0
        self.tStart = 0
        self.tEnd = 0
        self.bTimerOn = False


# --------------------------------------------------------------------------------------------------------------
# JSON CLASS : Timer manager
class cTimerList():
    def __init__(self):
        self.tList = dict()
    
    def TimerStart(self, tName: str):
        if tName not in self.tList:
            t = cTimer()
            self.tList[tName] = t
        
        self.tList[tName].tStart = time()
        self.tList[tName].bTimerOn = True
    
    def TimerEnd(self, tName: str):
        if tName not in self.tList:
            self.TimerStart(tName)
        
        if self.tList[tName].bTimerOn:
            self.tList[tName].tEnd = time()
            self.tList[tName].tTotal = self.tList[tName].tEnd - self.tList[tName].tStart
            self.tList[tName].bTimerOn = False
            
            return self.tList[tName].tTotal
        
        return 0
    
    def GetTime(self, tName: str):
        if tName in self.tList:
            return self.tList[tName].tTotal
        
        return 0
    
    def LogPrint(self, tName: str):
        if tName in self.tList:
            print(tName, "{:.3f}s".format(self.tList[tName].tTotal))
    
    def LogPrintAll(self):
        for tName in self.tList:
            print(tName + ":", "{:.3f}s".format(self.tList[tName].tTotal))
    
    def StrPrintAll(self):
        res = ""
        bFirst = True
        for tName in self.tList:
            if bFirst:
                res = tName + " : " + "{:.3f}s".format(self.tList[tName].tTotal)
                bFirst = False
            else:
                res = res + "\n" + tName + " : " + "{:.3f}s".format(self.tList[tName].tTotal)
        
        return res


# --------------------------------------------------------------------------------------------------------------
# JSON CLASS : Localtext list manage
class cLocalTextList:
    def __init__(self):
        self.dLocalText = dict()
        self.dKeyList = dict()
    
    def Add(self, lang: str, key: str, text: str, sheetName: str):
        dLang = dict()
        if lang in self.dLocalText:
            dLang = self.dLocalText[lang]
        
        if key in dLang:
            return "DUPLICATE_LOCAL_TEXT_KEY"

        dLang[key] = text
        self.dLocalText[lang] = dLang
        self.dKeyList[key] = sheetName
        
        return "PASS"


# --------------------------------------------------------------------------------------------------------------
# JSON CLASS : Reference data manager
class cReferenceDataList():
    def __init__(self):
        self.dOrigin = dict()
    
    def AddOrigin(self, referName: str, value, linkedData):
        referName = referName.lower()
        if referName in self.dOrigin:
            if len(linkedData) > 0:
                self.dOrigin[referName][value] = linkedData
            else:
                self.dOrigin[referName][value] = value
        else:
            dOriginMember = dict()
            if len(linkedData) > 0:
                dOriginMember[value] = linkedData
            else:
                dOriginMember[value] = value
            self.dOrigin[referName] = dOriginMember
    
    def Exists(self, referName: str, value):
        referName = referName.lower()
        if referName in self.dOrigin:
            if value in self.dOrigin[referName]:
                return "TRUE"
            else:
                return "FALSE"
        
        return "NO_REFERENCE"


# --------------------------------------------------------------------------------------------------------------
# JSON CLASS : Merge table data with original excel file for log link
class cMergeTableData:
    def __init__(self):
        self.fullName = ""
        self.fileName = ""
        self.sheetName = ""
        self.dataName = ""
        self.y = 0
        self.x = 0
        self.value = ""

    def SetData(self, orgFullName: str, orgFileName: str, orgSheetName: str, orgDataName: str, orgY: int, orgX: int, value):
        self.fullName = orgFullName
        self.fileName = orgFileName
        self.sheetName = orgSheetName
        self.dataName = orgDataName
        self.y = orgY
        self.x = orgX
        self.value = value


# --------------------------------------------------------------------------------------------------------------
# JSON CLASS : Merge table manager
class cMergeTable:
    def __init__(self):
        self.mergeName = ""
        self.mergeGroup = ""
        self.mergeCount = 0
        
        self.dSheet = dict()
        self.dTS = dict()
        self.dSD = dict()
        
        self.dTitle = dict()
        self.dOldTitlePos = dict()
        self.dUID = dict()
        self.dGroupID = dict()
        
        self.ts = cTableSchema()
        self.sd = cSheetData()
        
        self.data = dict()
    
    def AddSheet(self, ts: cTableSchema, sd: cSheetData):
        if ts.sheetName not in self.dSheet:
            self.dSheet[ts.sheetName] = len(self.dSheet) + 1
            self.dTS[ts.sheetName] = ts
            self.dSD[ts.sheetName] = sd
            
    # def DataAdd(self, y: int, x: int, value: cMergeTableData):
    #     if y in self.data:
    #         self.data[y][x] = value
    #     else:
    #         values = dict()
    #         values[x] = value
    #         self.data[y] = values
    
    def DataFind(self, y: int, x: int):
        res = cMergeTableData()
        
        if y in self.data:
            if x in self.data:
                res = self.data[y][x]
            
        return res


# --------------------------------------------------------------------------------------------------------------
# JSON CLASS : Merge table index split key manager
class cIndex:
    def __init__(self):
        self.idx = dict()
        self.idxDataLine = dict()        
        self.yIndex = dict()
        
        self.cntIdx = 1
        self.cntN = 0
        self.cntMax = 30

    def Init(self, mergeCount: int):
        self.cntMax = mergeCount
    
    def AddIndexKey(self, indexKey: int, y: int):
        if int(indexKey) > 0:
            if indexKey not in self.idx:
                self.idx[indexKey] = self.cntIdx
                
                if self.cntN >= self.cntMax:
                    self.cntIdx = self.cntIdx + 1
                    self.cntN = 1
                else:
                    self.cntN = self.cntN + 1
                                
            if y not in self.yIndex:
                self.yIndex[y] = self.cntIdx
                if self.cntIdx in self.idxDataLine:
                    self.idxDataLine[self.cntIdx].append(y)
                else:
                    nList = list()
                    nList.append(y)
                    self.idxDataLine[self.cntIdx] = nList
    
    def GetIndexNoByIndexKey(self, indexKey):
        if indexKey in self.idx:
            return self.idx[indexKey]
        return -1
    
    def GetIndexKeybyIndexNo(self, idxNo: int):
        res = dict()
        for tKey in self.idx:
            if idxNo == self.idx[tKey]:
                res[tKey] = idxNo
        return res
    
    def GetIndexMax(self):
        return self.cntIdx    


# --------------------------------------------------------------------------------------------------------------
# JSON CLASS : Main parameters for json export
class cParam:
    def __init__(self, uiWin):
        self.uiWin = uiWin
        self.cPath = os.getcwd()
        self.cu = 0
        if cpu_count() >= 10:
            self.cu = int(cpu_count() / 2.0) + 1
        else:
            self.cu = 3

        self.cu = min(self.cu, 6)

        self.dMergeList = dict()
        self.listSheet = dict()
        self.listLog = cLogList()
                
        self.dbEnum = cEnumList()
        self.dbText = cLocalTextList()
        self.dbMerge = dict()
                
        self.dbData = dict()
        self.dbChangeValue = dict()
        self.dbOrigin = cReferenceDataList()
        self.dbArraySize = cArraySizeRefCheck()
        
        self.pathOutput = self.cPath + _SEP + "output" + _SEP
        self.pathError = self.cPath  + _SEP + "error" + _SEP
        self.errorFileName = ""
        
        self.bEnum = False
        self.bText = False
        self.lMerge = list()
        self.bData = False
        
        self.bNeedEnum = False
        self.bNeedText = False
        self.lNeedMerge = list()        
        self.refExcel = list()
        
        self.tl = cTimerList()

        self.pathOutputJson = self.pathOutput + self.uiWin.strTablePath.replace("/",_SEP) #"Application\\Bundles\\Tables\\Generated\\"
        self.pathOutputScript = self.pathOutput + self.uiWin.strScriptPath.replace("/",_SEP) #"Application\\Scripts\\Tables\\Generated\\"

        os.makedirs(os.path.dirname(self.pathOutput), exist_ok = True)
        os.makedirs(os.path.dirname(self.pathError), exist_ok = True)

        os.makedirs(os.path.dirname(self.pathOutputJson), exist_ok = True)
        os.makedirs(os.path.dirname(self.pathOutputScript), exist_ok = True)

# --------------------------------------------------------------------------------------------------------------
# Global Param
p = None #cParam()


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Main function json export
def ExportJsonMain(uiWin):
    global p

    if uiWin.state == MainUIState.DISABLED:
        return

    uiWin.ButtonDisable()

    if uiWin.autoSaveToProjectVar.get() == 1:
        if os.path.isdir(uiWin.strProjectFullPath):
            strProjectName = os.path.basename(os.path.normpath(uiWin.strProjectFullPath))
            if sys.platform != "darwin" and find_unity_editor(strProjectName) == None:
                messagebox.showerror("Export", f"{strProjectName} UnityEditor가 실행되고 있지 않습니다.\n{strProjectName} UnityEditor를 실행해 주세요.")
                uiWin.ButtonEnable()
                return
        else:
            messagebox.showerror("Export", "ProjectFullPath경로가 올바른 UnityProject경로가 아닙니다.\n세팅에서 ProjectFullPath경로를 올바른 UnityProject경로로 설정해주세요.")
            uiWin.ButtonEnable()
            return


    p = cParam(uiWin)
    p.dMergeList = uiWin.dMergeList

    p.tl.TimerStart("[Run] Script")
    pBar = cProgressPopup(uiWin)
    maxRate = 5
    pBar.SetInit("Exporting checked excel files", "Prepare export process", "0 / " + str(maxRate), 0)    

    p.tl.TimerStart("[Check] Export File List")
    res = CheckExportFileList(uiWin, pBar)
    p.tl.TimerEnd("[Check] Export File List")

    if res == True:
        # Read checked excel files
        p.tl.TimerStart("[Read] Excel Files")
        ReadExcelFiles(True, uiWin, pBar)
        p.tl.TimerEnd("[Read] Excel Files")
        
        # Read reference excel files
        if p.listLog.LogCount() == 0:            
            if p.bNeedEnum is True and p.bEnum is False:
                for fileName in uiWin.dExcelList:
                    if uiWin.dExcelList[fileName]["type"] == "TYPE" and fileName not in p.refExcel:
                        p.refExcel.append(fileName)
                    
            if p.bNeedText is True and p.bText is False:
                for fileName in uiWin.dExcelList:
                    if uiWin.dExcelList[fileName]["type"] == "LOCALTEXT" and fileName not in p.refExcel:
                        p.refExcel.append(fileName)

            for mergeName in p.lNeedMerge:
                for fileName in uiWin.dMergeList[mergeName]:
                    if uiWin.dExcelList[fileName]["type"] == "MERGE" and fileName not in p.refExcel and mergeName not in p.lMerge:
                        p.refExcel.append(fileName)
                        
            if len(p.refExcel) > 0:
                p.tl.TimerStart("[Ref Read] Read Reference Excel Files")
                ReadExcelFiles(False, uiWin, pBar)
                p.tl.TimerEnd("[Ref Read] Read Reference Excel Files")
            
        # Merge Table
        if p.listLog.LogCount() == 0:
            for mergeName in p.lMerge:
                p.tl.TimerStart("[Table Merge] " + mergeName)
                MergeTables(pBar, mergeName)
                p.tl.TimerEnd("[Table Merge] " + mergeName)
            
            for mergeName in p.lNeedMerge:
                p.tl.TimerStart("[Table Merge] " + mergeName)
                MergeTables(pBar, mergeName)
                p.tl.TimerEnd("[Table Merge] " + mergeName)

        # Check Data Validation
        if p.listLog.LogCount() == 0:
            p.tl.TimerStart("[Check] Data Validation Check")
            DataVaildCheck(pBar)
            p.tl.TimerEnd("[Check] Data Validation Check")

        # Make Json Files
        if p.listLog.LogCount() == 0:
            p.tl.TimerStart("[Json] Generate Json Files")
            try:
                delete_folder_contents(p.pathOutput)
            except FileDeletionError as e:
                pBar.Exit()
                messagebox.showinfo("Json export failed", e)
                uiWin.ButtonEnable()
                return

            os.makedirs(os.path.dirname(p.pathOutputJson), exist_ok = True)
            os.makedirs(os.path.dirname(p.pathOutputScript), exist_ok = True)

            MakeJsonFiles(pBar)
            p.tl.TimerEnd("[Json] Generate Json Files")
            p.tl.TimerEnd("[Run] Script")
            pBar.Exit()
            messagebox.showinfo("Json export completed", p.tl.StrPrintAll())
            if uiWin.autoSaveToProjectVar.get() == 1 :
                if is_unity_project(uiWin.strProjectFullPath) :
                    try:
                        copy_folder(p.pathOutput, uiWin.strProjectFullPath + "/tmp/")
                    except Exception  as e:
                        messagebox.showerror("Export", f"{uiWin.strProjectFullPath}로 익스포트된 파일 복사에 실패하였습니다." + e)
                        uiWin.ButtonEnable()
                        return
                    strProjectName = os.path.basename(os.path.normpath(uiWin.strProjectFullPath))
                    messagebox.showinfo("Export", f"{strProjectName}로 익스포트된 파일 복사에 성공하였습니다.\n확인을 눌러 에셋생성을 진행해주세요" )
                    focus_unity_editor((os.path.basename(os.path.normpath(uiWin.strProjectFullPath))))
                else:
                    messagebox.showerror("Export", "ProjectFullPath경로가 올바른 UnityProject경로가 아닙니다.\n세팅에서 ProjectFullPath경로를 올바른 UnityProject경로로 설정해주세요.")
        else:
            SaveErrorLogFile(pBar)
            p.tl.TimerEnd("[Run] Script")            
            pBar.Exit()
            messagebox.showwarning("Json export failed", f"{str(p.listLog.LogCount())} error logs ( {p.errorFileName} )")
        
    else:
        p.tl.TimerEnd("[Run] Script")
        pBar.Exit()
        messagebox.showwarning("Export", "No checked file")
    
    uiWin.ButtonEnable()

class FileDeletionError(Exception):
    pass

def delete_folder_contents(folder_path):
    # 하위 폴더와 파일을 재귀적으로 삭제
    for item in os.listdir(folder_path):
        item_path = os.path.join(folder_path, item)

        try:
            if os.path.isfile(item_path):
                os.remove(item_path)
            elif os.path.isdir(item_path):
                delete_folder_contents(item_path)  # 하위 폴더 재귀 호출
                os.rmdir(item_path)  # 폴더 삭제
        except Exception as e:
            raise FileDeletionError(f"삭제 실패: {item_path} - 오류: {e}")
# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Prepare Export
def CheckExportFileList(uiWin, pBar):
    global p
    
    pBar.SetText("Check Export File List", "")
    pBar.SetRate(0)
    
    if len(uiWin.dCheckedFileList) == 0:
        return False
    
    for fileName in uiWin.dCheckedFileList:
        dataType = uiWin.dExcelList[fileName]["type"]
        match(dataType):
            case "TYPE":
                p.bEnum = True
            case "LOCALTEXT":
                p.bText = True
            case "MERGE":
                mergeName = uiWin.dExcelList[fileName]["mergeName"]
                if mergeName not in p.lMerge:
                    p.lMerge.append(mergeName)
            case "DATA":
                p.bData = True
    
    for fileName in uiWin.dExcelList:
        dataType = uiWin.dExcelList[fileName]["type"]
        match(dataType):
            case "TYPE":
                if p.bEnum == True and fileName not in uiWin.dCheckedFileList:
                    uiWin.dCheckedFileList.append(fileName)
            case "LOCALTEXT":
                if p.bText == True and fileName not in uiWin.dCheckedFileList:
                    uiWin.dCheckedFileList.append(fileName)
            case "MERGE":
                mergeName = uiWin.dExcelList[fileName]["mergeName"]
                if mergeName in p.lMerge and fileName not in uiWin.dCheckedFileList:
                    uiWin.dCheckedFileList.append(fileName)
            case "DATA":
                pass
    
    uiWin.RefreshListExcel()
    
    pBar.SetRate(100)
    return True


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Read Excel Files (Excute Multiprocess)
def ReadExcelFiles(bChecked, uiWin, pBar):
    global p

    paramList = list()

    # Prepare multiprocessing parameters
    if bChecked:
        readFileList = uiWin.dCheckedFileList
        pBar.strTitle = "Read Excel Files"
    else:
        readFileList = p.refExcel
        pBar.strTitle = "Read Reference Excel Files"

    pBar.SetText(pBar.strTitle, "")
    pBar.SetRate(0)
    
    for fileName in readFileList:
        fullPath = uiWin.dExcelList[fileName]["fullPath"]
        path = Path(fullPath)
        if path.is_file():
            l = [str(path.parent), fileName, bChecked, p.dMergeList]
            paramList.append(l)

    pBar.t_count = len(paramList)
    pBar.count = 0
    res = list()

    # Read files with multiprocessing
    if pBar.t_count > 0:
        mt = Pool(min(p.cu, len(paramList)))
        mt_res = mt.map_async(ReadExcel, paramList, chunksize=1)
        
        while mt_res.ready() == False:
            pBar.count = sum(x is not None for x in mt_res._value)
            pBar.SetText(pBar.strTitle, f"{pBar.count} / {pBar.t_count}")
            pBar.SetRate(pBar.count / pBar.t_count * 100)
            sleep(0.25)
        res = mt_res.get()
        mt.close()
        mt.join()

    pBar.SetText(pBar.strTitle, f"{pBar.count} / {pBar.t_count}")
    pBar.SetRate(pBar.count / pBar.t_count * 100)
    
    # Remove empty result (excel file with no table sheet)
    for resExcel in res.copy():
        if resExcel == []:
            res.remove(resExcel)
    
    # Merge result
    # resSheet[0] = data type
    # resSheet[1] = sheetName, ts, sd, bChecked, dbArraySize, dbOrigin, logs
    for resExcel in res:
        for resSheet in resExcel:
            ts = resSheet[1][1]
            sd = resSheet[1][2]
            if resSheet[1][0] in p.listSheet:
                p.listLog.Add(ts.fullPath, ts.fileName, ts.sheetName, ts.sheetName, 1, 1, "테이블 이름이 중복됐습니다 (" + ts.sheetName + ")")
            else:
                p.listSheet[resSheet[1][0]] = ts.fileName        
                
                if bChecked:
                    if ts.bNeedEnum: p.bNeedEnum = True
                    if ts.bNeedText: p.bNeedText = True
                    if len(ts.bNeedMerge)>0:
                        for mergeName in ts.bNeedMerge:
                            if mergeName not in p.lNeedMerge and mergeName not in p.lMerge:
                                p.lNeedMerge.append(mergeName)
                    
                    for refSheetName in ts.refSheet:
                        if refSheetName in uiWin.dSheetList:                                    
                            fileName = uiWin.dSheetList[refSheetName]
                            if refSheetName not in p.listSheet and fileName not in uiWin.dCheckedFileList and fileName not in p.refExcel:
                                p.refExcel.append(fileName)
                        else:
                            p.listLog.Add(ts.fullPath, ts.fileName, ts.sheetName, ts.sheetName, 1, 1, "존재하지 않는 테이블을 참조하고 있습니다 (" + refSheetName + ")")
                
                MergeArraySize(resSheet[1][4])
                MergeOrigin(resSheet[1][5])
                MergeLog(resSheet[1][6])
                
                match resSheet[0]:
                    case "#type":
                        GenerateEnumDB(ts, sd)
                        SaveTableData(ts, sd, "type")
                    case "#localtext":
                        GenerateTextDB(ts, sd)
                        SaveTableData(ts, sd, "localtext")
                    case "#merge":
                        mergeName = ts.mergeData[0]
                        mergeGroup = ts.mergeData[1]
                        mergeCount = ts.mergeData[2]
                        
                        GenerateMergeDB(ts, sd, mergeName, mergeGroup, mergeCount)
                        SaveTableData(ts, sd, mergeName)
                    case "#data":
                        if resSheet[1][3] == True:
                            SaveTableData(ts, sd, "data")
                        else:
                            SaveTableData(ts, sd, "ref")        


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Read Excel File
def ReadExcel(paramList: list):
    fullPath = paramList[0]
    fileName = paramList[1]
    bChecked = paramList[2]
    dMergeList = paramList[3]

    wb = load_workbook(filename = fullPath + _SEP + fileName, read_only = True, data_only = True)
    res = list()
        
    for ws in wb.worksheets:
        if ws.max_row > 2:            
            match(ws.cell(2,1).value):
                case "#type" | "#localtext" | "#merge" | "#data":
                    sheet_res = list()
                    sheet_res.append(ws.cell(2,1).value)
                    ws_list = np.array([[cell.value for cell in row] for row in ws.iter_rows()])                    
                    sheet_res.append(ReadSheet(ws_list, fullPath, fileName, ws.title, bChecked, dMergeList))
                    res.append(sheet_res)

    return res


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Read Sheet
def ReadSheet(ws, fullPath: str, fileName: str, sheetName: str, bChecked: bool, dMergeList: dict):    
    ts = cTableSchema()
    ts.Init(ws, fullPath, fileName, sheetName, dMergeList)
    
    posKey = ts.posKey
    keyList = dict()
    sd = cSheetData()
    dbArraySize = cArraySizeRefCheck()
    dbOrigin = cReferenceDataList()    
    logs = cLogList()

    if len(ts.logList) > 0:
        for logText in ts.logList:
            val = ParseToList(logText, ",")
            logPos = int(val[1])
            match(val[0]):
                case "NO_CHECKED":
                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[logPos], 1, logPos, "추출할 데이터 필드가 없습니다")
                case "NO_KEY_FIELD":
                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[logPos], 2, logPos, "키 필드가 정의되어 있지 않습니다")
                case "NOT_SUB_FOLDER":
                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[logPos], 0, 0, "서브 폴더가 올바른 규칙이 아닙니다")
                case "SPACE_IN_TITLE":
                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[logPos], 1, logPos, "필드 이름에 빈 칸이 있습니다")
                case "DUPLICATE_TITLE":
                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[logPos], 1, logPos, "필드 이름이 중복 사용되었습니다")
                case "NO_DATA_TYPE":
                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[logPos], 2, logPos, "데이터 타입이 정의되어 있지 않습니다")
                case "ENUM_NULL_TO_SPACE":
                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[logPos], 2, logPos, "ENUM 타입 필드는 NULL 변환을 사용할 수 없습니다")
                case "EMPTY_KEY":
                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[logPos], 2, logPos, "키 필드는 NULL 변환을 사용할 수 없습니다")
                case "EMPTY_REFERECNE_VALUE":
                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[logPos], 2, logPos, "참조되는 필드는 NULL 변환을 사용할 수 없습니다")
                case "EMPTY_UNIQUE_VALUE":
                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[logPos], 2, logPos, "고유 값을 사용하는 필드는 NULL 변환을 사용할 수 없습니다")
                case "UNVALID_REFRENCE_FIELD":
                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[logPos], 2, logPos, "enumref에서 존재하지 않는 필드를 참조하고 있습니다")
                case "INVALID_REF_NAME":
                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[logPos], 2, logPos, "check에서 참조하는 테이블의 이름과 데이터의 이름이 정확하지 않습니다")
                case "INVALID_CHANGE_FILED":
                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[logPos], 2, logPos, "change에서 존재하지 않는 필드를 참조하고 있습니다")
                case "INVALID_ARRAYREF_NAME":
                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[logPos], 2, logPos, "arraysizeref에서 존재하지 않는 필드를 참조하고 있습니다")
                case "INVALID_ARRAYCNTREF_NAME":
                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[logPos], 2, logPos, "arraysizecnt에서 존재하지 않는 필드를 참조하고 있습니다")
    else:
        for y in range(4, ts.maxRow):
            if ws[y][posKey] is None:
                keyValue = ""
            else:
                keyValue = str(ws[y][posKey])
            
            if len(keyValue) > 0 or ts.bValueOnly :
                if keyValue in keyList and not ts.bValueOnly :
                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[posKey], y, posKey, "키 값이 중복됐습니다 (" + keyValue + ")")
                else:
                    if ts.bValueOnly :
                       keyList[keyValue] = y
                    dRow = dict()
                    
                    for x in range(1, ts.maxCol):
                        if x in ts.cExportX or x in ts.cCheckX:
                            # value = None Check
                            if (ws[y][x]) is None:
                                value = ""
                            else:
                                value = ws[y][x]
                            
                            # Float Accuracy
                            if x not in ts.dArray:
                                match(ts.dDataType[x]):
                                    case 'float' | 'int' | 'long' | 'double':
                                        try:
                                            value = round(float(value), 8)
                                            if int(value) == value:
                                                value = str(int(value))
                                            else:
                                                value = str(value)
                                        except ValueError:
                                            value = str(value)
                                    case _:
                                        value = str(value)
                            else:
                                value = str(value)

                            # Null Check
                            if value.lower() == "null":
                                if x not in ts.dNull:
                                    logs.Add(fullPath, fileName, sheetName, ts.hdTitle[x], y, x, "NULL 선언이 되지 않은 필드입니다")
                                else:
                                    value = ""
                        
                            # New Line Check
                            if '\r' in value or '\n' in value:
                                logs.Add(fullPath, fileName, sheetName, ts.hdTitle[x], y, x, "텍스트에 개행 문자가 포함되어 있습니다")

                            tList = list()
                            if len(value) > 0:
                                if x in ts.dArray:
                                    tList = ParseToList(value, ",")
                                else:
                                    tList.append(value)
                        
                            # Add Prefix / Suffix
                            if len(value) > 0 and (x in ts.dPrefix or x in ts.dSuffix):
                                nValue = ""
                                
                                bFirst = True
                                for tValue in tList:
                                    if bFirst:
                                        bFirst = False
                                    else:
                                        nValue = nValue + ","
                                    
                                    if len(tValue) > 0:
                                        if x in ts.dPrefix: nValue = nValue + ts.dPrefix(x)
                                        nValue = nValue + tValue
                                        if x in ts.dSuffix: nValue = nValue + ts.dSuffix(x)
                                value = nValue
                                                
                            # Data Type Check
                            res = "PASS"
                            if x in ts.dArray:                        
                                if len(value) > 0:
                                    for tValue in tList:
                                        res = TypeCheck(ts.dDataType[x], str(tValue))
                                        if res != "PASS": break
                                
                                if x in ts.dArraySize and len(tList) > 0:
                                    if str(len(tList)) != ts.dArraySize[x]:
                                        logs.Add(fullPath, fileName, sheetName, ts.hdTitle[x], y, x, "배열의 개수가 지정된 값과 다릅니다 (" + value + ")")
                                
                                if x in ts.dArraySizeOrg or x in ts.dArraySizeRef:
                                    dbArraySize.AddArraySize(ts.sheetName, y, x, len(tList))
                            else:
                                res = TypeCheck(ts.dDataType[x], value)
                            
                            # Array Size Check
                            if x in ts.dArraySizeCnt:
                                try:
                                    arrSize = int(value)
                                except ValueError:
                                    arrSize = 0
                                dbArraySize.AddArraySize(ts.sheetName, y, x, arrSize)
                            
                            # Type Error Log
                            match res:
                                case "INT_VALUE_ERROR": logs.Add(fullPath, fileName, sheetName, ts.hdTitle[x], y, x, "int 타입의 값이 아닙니다 (" + value + ")")
                                case "INT_VALUE_NULL": logs.Add(fullPath, fileName, sheetName, ts.hdTitle[x], y, x, "int 타입에 빈 값이 입력되어 있습니다")
                                case "FLOAT_VALUE_ERROR": logs.Add(fullPath, fileName, sheetName, ts.hdTitle[x], y, x, "float 타입의 값이 아닙니다 (" + value + ")")
                                case "FLOAT_VALUE_NULL": logs.Add(fullPath, fileName, sheetName, ts.hdTitle[x], y, x, "float 타입에 빈 값이 입력되어 있습니다")
                                case "BOOL_VALUE_ERROR": logs.Add(fullPath, fileName, sheetName, ts.hdTitle[x], y, x, "bool 타입의 값이 아닙니다 (" + value + ")")
                                case "BOOL_VALUE_NULL": logs.Add(fullPath, fileName, sheetName, ts.hdTitle[x], y, x, "bool 타입에 빈 값이 입력되어 있습니다")
                                case "TYPE_VALUE_SPCAE_ERROR": logs.Add(fullPath, fileName, sheetName, ts.hdTitle[x], y, x, "타입에 공백이 포함되어 있습니다 (" + value + ")")
                                case "TYPE_VALUE_NULL": logs.Add(fullPath, fileName, sheetName, ts.hdTitle[x], y, x, "타입에 빈 값이 입력되어 있습니다")
                                case "STRING_VALUE_QUOTE_ERROR": logs.Add(fullPath, fileName, sheetName, ts.hdTitle[x], y, x, "문자열에 \"가 포함되어 있습니다")
                            
                            match ts.dDataType[x]:
                                case "bool":
                                    if value.lower() == "true":
                                        value = "true"
                                    if value.lower() == "false":
                                        value = "false"
                            
                            dRow[x] = str(value)
                
                    sd.data[y] = dRow
                    
                    for x in ts.dOrigin:                
                        if x in ts.dChange:
                            dbOrigin.AddOrigin(ts.dOrigin[x], dRow[x], sd.Find(y, ts.dChange[x]))
                        else:
                            dbOrigin.AddOrigin(ts.dOrigin[x], dRow[x], "")

                    if ts.bValueOnly :
                        break
    return ts.sheetName, ts, sd, bChecked, dbArraySize, dbOrigin, logs


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Generate Text List by Lang Type (from Multirpocess result)
def GenerateTextDB(ts: cTableSchema, sd: cSheetData):
    global p
    
    for y in sd.data:
        keyValue = sd.data[y][ts.posKey]
        
        if keyValue not in p.dbText.dKeyList:
            for x in sd.data[y]:
                if x in ts.cExportX and x != ts.posKey:
                    res = p.dbText.Add(ts.hdTitle[x], keyValue, str(sd.data[y][x]), ts.sheetName)
                    if res != "PASS":
                        p.listLog.Add(ts.fullPath, ts.fileName, ts.sheetName, "Local Text Key", y, ts.posKey, "로컬 텍스트의 키가 중복 사용됐습니다. (" + str(keyValue) + ")")
        else:
            p.listLog.Add(ts.fullPath, ts.fileName, ts.sheetName, "Local Text Key", y, ts.posKey, "로컬 텍스트의 키가 중복 사용됐습니다. (" + str(keyValue) + ")")


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Generate Type List (from Multirpocess result)
def GenerateEnumDB(ts: cTableSchema, sd: cSheetData):    
    global p
    
    posGroup = 0
    posText = 0
    posValue = 0
    posComment = 0
    
    for x in range(1, ts.maxCol+1):
        if x in ts.hdTitle:
            match ts.hdTitle[x]:
                case "type_group": posGroup = x
                case "type_text": posText = x
                case "type_value": posValue = x
                case "type_desc": posComment = x
    
    if posGroup == 0 or posText == 0 or posValue == 0 or posComment == 0:
        p.listLog.Add(ts.fullPath, ts.fileName, ts.sheetName, "Type 테이블 오류", 1, 1, "Type 테이블의 구조가 규칙과 다릅니다")
    else:
        for y in sd.data:
            sEnum = sd.data[y][posGroup]
            sText = sd.data[y][posText]
            sValue = sd.data[y][posValue]
            sComment = sd.data[y][posComment]
            
            res = p.dbEnum.AddType(ts.sheetName, sEnum, sText, sValue, sComment, ts.bNocode)
            
            match res:
                case "NULL_TYPE_GROUP": p.listLog.Add(ts.fullPath, ts.fileName, ts.sheetName, "Type Group", y, posText, "타입 그룹 값이 비어있습니다")
                case "NULL_TYPE_TEXT": p.listLog.Add(ts.fullPath, ts.fileName, ts.sheetName, "Type Text", y, posText, "타입 텍스트 값이 비어있습니다")
                case "NULL_TYPE_VALUE": p.listLog.Add(ts.fullPath, ts.fileName, ts.sheetName, "Type Value", y, posText, "타입 밸류 값이 비어있습니다")
                case "DUPLICATE_TYPE_GROUP_ANOTHER_SHEET": p.listLog.Add(ts.fullPath, ts.fileName, ts.sheetName, "Type Group", y, posText, "타입 그룹 값이 다른 시트에서 사용 중입니다 (" + sEnum + ")")
                case "DUPLICATE_TYPE_TEXT_IN_TYPE_GROUP": p.listLog.Add(ts.fullPath, ts.fileName, ts.sheetName, "Type Text", y, posText, "타입 텍스트가 중복됐습니다 (" + sEnum + ":" + sText + ")")
                case "DUPLICATE_TYPE_VALUE_IN_TYPE_GROUP": p.listLog.Add(ts.fullPath, ts.fileName, ts.sheetName, "Type Value", y, posText, "타입 밸루가 중복됐습니다 (" + sEnum + ":" + str(sValue) + ")")


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Generate Merge Table DB by mergeName (from Multirpocess result)
def GenerateMergeDB(ts: cTableSchema, sd: cSheetData, mergeName: str, mergeGroup: str, mergeCount: int):
    global p
    
    mergeTable = cMergeTable()
    if mergeName in p.dbMerge:
        mergeTable = p.dbMerge[mergeName]
    
    mergeTable.mergeName = mergeName
    mergeTable.mergeGroup = mergeGroup
    mergeTable.mergeCount = mergeCount    
    mergeTable.AddSheet(ts, sd)
    
    p.dbMerge[mergeName] = mergeTable


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Data Vaild Check Manage (Excute Multiprocess)
def DataVaildCheck(pBar):
    global p
    
    paramList = list()
    pBar.strTitle = "Check Data Validation"
    pBar.SetText(pBar.strTitle, "")
    pBar.SetRate(0)
    
    if "data" in p.dbData:
        tData = p.dbData["data"]
        
        for tKey in tData:
            ts = tData[tKey]["ts"]
            sd = tData[tKey]["sd"]
            l = [ts, sd, ts.sheetName, p.dbEnum, p.dbText, p.dbArraySize, p.dbOrigin, False, cMergeTable()]
            paramList.append(l)
    
    for mergeName in p.lMerge:
        mergeTable = p.dbMerge[mergeName]
        ts = mergeTable.ts
        sd = mergeTable.sd
        l = [ts, sd, mergeName, p.dbEnum, p.dbText, p.dbArraySize, p.dbOrigin, True, mergeTable]
        paramList.append(l)
    
    pBar.t_count = len(paramList)
    pBar.count = 0
    res = list()
    
    if pBar.t_count > 0:
        mt = Pool(min(p.cu, len(paramList)))
        mt_ret = mt.map_async(DataValidCheckTable, paramList, chunksize=1)
        
        pBar.count = 0
        while mt_ret.ready() == False:
            pBar.count = sum(x is not None for x in mt_ret._value)
            pBar.SetText(pBar.strTitle, f"{pBar.count} / {pBar.t_count}")
            pBar.SetRate(pBar.count / pBar.t_count * 100)
            sleep(0.5)
        
        res = mt_ret.get()
        mt.close()
        mt.join()

        pBar.SetText(pBar.strTitle, f"{pBar.count} / {pBar.t_count}")
        pBar.SetRate(pBar.count / pBar.t_count * 100)
        
        for resTable in res.copy():
            if resTable == []:
                res.remove(resTable)
        
        for resTable in res:
            SaveTableData(resTable[1], resTable[2], "data")
            MergeLog(resTable[3])        
            MergeChangeValue(resTable[4])
        
        pBar.SetText(pBar.strTitle, f"{pBar.count} / {pBar.t_count}")
        pBar.SetRate(pBar.count / pBar.t_count * 100)


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Check data validation
def DataValidCheckTable(paramList: list):    
    ts = paramList[0]
    sd = paramList[1]
    sheetName = paramList[2]
    dbEnum = paramList[3]
    dbText = paramList[4]
    dbArraySize = paramList[5]
    dbOrigin = paramList[6]
    mergeFlag = paramList[7]
    mergeTable = paramList[8]
    
    return(CommonDataCheck(ts, sd, sheetName, dbEnum, dbText, dbArraySize, dbOrigin, mergeFlag, mergeTable))


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Common Data Valid Check Function
def CommonDataCheck(ts: cTableSchema, sd: cSheetData, sheetName: str, dbEnum: cEnumList, dbText: cLocalTextList, dbArraySize: cArraySizeRefCheck, dbOrigin: cReferenceDataList, mergeFlag: bool, mergeTable: cMergeTable):
    dUniqueTest = dict()
    dbChangeValue = dict()
    logs = cLogList()
    
    fileName = ts.fileName
    fullPath = ts.fullPath
    orgSheetName = sheetName
    orgY = 0
    orgX = 0
    
    for y in sd.data:
        dRow = sd.data[y]
        
        for x in dRow:
            orgY = y
            orgX = x
            tValue = dRow[x]
            
            if mergeFlag == True:
                mergeOrgData = mergeTable.DataFind(y, x)
                fileName = mergeOrgData.fileName
                fullPath = mergeOrgData.fullName
                orgSheetName = mergeOrgData.sheetName
                orgY = mergeOrgData.y
                orgX = mergeOrgData.x
            
            # Unique Value Check
            if str(tValue) != "null":
                if x in ts.dUnique:
                    tKey = str(x) + "/" + str(tValue)
                    if tKey in dUniqueTest:
                        logs.Add(fullPath, fileName, orgSheetName, ts.hdTitle[x], orgY, orgX, "중복된 값이 입력됐습니다 (" + str(tValue) + ")")
                    else:
                        dUniqueTest[tKey]  = x
            
            # Type Check
            if x in ts.dEnum:
                sType = ts.dEnum[x].upper()
                
                if sType not in dbEnum.dEnumList and x not in ts.dEnumRef:
                    logs.Add(fullPath, fileName, orgSheetName, ts.hdTitle[x], orgY, orgX, "참조하는 타입이 정의되어 있지 않습니다. (" + str(ts.dEnum[x]) + ")")
                else:
                    if x not in ts.dChanged:
                        ts.dChanged[x] = ts.hdTitle[x]
                    
                    if x in ts.dEnumRef:
                        sType = dRow[ts.dEnumRef[x]].upper()
                    
                    #if sType == "NONE" and x in ts.dArray and len(tValue) == 0:
                    if sType == "NONE" and x in ts.dArray and len(tValue) == 0:
                        dbChangeValue[sheetName + "/" + str(y) + "/" + str(x)] = ""
                    elif sType == "NONE":
                        if tValue == "NONE":
                            dbChangeValue[sheetName + "/" + str(y) + "/" + str(x)] = 0
                        else:
                            logs.Add(fullPath, fileName, orgSheetName, ts.hdTitle[x], orgY, orgX, "참조하는 타입이 없을 경우 값을 NONE으로 입력해야 합니다 (" + str(tValue) + ")")
                    elif sType in dbEnum.dEnumList:
                        if len(tValue) > 0:
                            if x in ts.dArray:
                                tList = ParseToList(tValue, ",", False, False, False)
                            else:
                                tList = [tValue]
                            
                            nList = list()
                            
                            for tKey in tList:
                                sText = str(tKey).upper()
                                if sType + "/" + sText in dbEnum.dTextCheck:
                                    nList.append(dbEnum.dTextCheck[sType + "/" + sText])
                                else:
                                    logs.Add(fullPath, fileName, orgSheetName, ts.hdTitle[x], orgY, orgX, "잘못된 타입 텍스트입니다 (" + sText + ")")
                            
                            nValue = ListToString(nList, ",")
                            dbChangeValue[sheetName + "/" + str(y) + "/" + str(x)] = nValue
                        else:
                            if x not in ts.dArray:
                                logs.Add(fullPath, fileName, orgSheetName, ts.hdTitle[x], orgY, orgX, "TYPE 필드에 입력된 값이 없습니다 (" + str(tValue) + ")")
                    else:
                        logs.Add(fullPath, fileName, orgSheetName, ts.hdTitle[x], orgY, orgX, "참조하는 TYPE이 정의되어 있지 않습니다 (" + str(sType) + "/" + str(tValue) + ")")

            if len(tValue) > 0:
                if x in ts.dArray:
                    tList = ParseToList(tValue, ",", False, False, False)
                else:
                    tList = [tValue]
    
            # Localtext Check
            if x in ts.dDataType:
                sType = ts.dDataType[x]
                if sType.lower() == "localtext":
                    if len(tValue) > 0:
                        for tKey in tList:
                            sText = str(tKey)
                            if sText not in dbText.dKeyList:
                                logs.Add(fullPath, fileName, orgSheetName, ts.hdTitle[x], orgY, orgX, "존재하지 않는 로컬텍스트 키를 참조하고 있습니다 (" + sText + ")")
            
            # Prepare Change Filed Value
            if x in ts.dChange:
                if x not in ts.dChanged:
                    ts.dChanged[x] = ts.hdTitle[x]
                dbChangeValue[sheetName + "/" + str(y) + "/" + str(x)] = dRow[ts.dChange[x]]
            
            # Array Size Check
            if x in ts.dArraySizeRef:
                arrSize = dbArraySize.GetSize(ts.sheetName, y, x)
                arrSizeOrg = dbArraySize.GetSize(ts.sheetName, y, ts.dArraySizeRef[x])
                if arrSize == -1:
                    logs.Add(fullPath, fileName, orgSheetName, ts.hdTitle[x], orgY, orgX, "배열 크기의 정보가 없습니다")
                elif arrSizeOrg == -1:
                    logs.Add(fullPath, fileName, orgSheetName, ts.hdTitle[x], orgY, orgX, "참조하는 항목의 배열 크기 정보가 없습니다")
                elif arrSize != arrSizeOrg:
                    logs.Add(fullPath, fileName, orgSheetName, ts.hdTitle[x], orgY, orgX, "두 필드의 배열 크기가 다릅니다 (" + str(arrSize) + " != " + str(arrSizeOrg) + ")")
            
            # Reference Value Check
            if x in ts.dValidCheck:
                sLink = ts.dValidCheck[x]
                if sLink in dbOrigin.dOrigin:
                    dLinkData = dbOrigin.dOrigin[sLink]
                    
                    for fKey in dLinkData:
                        if fKey != dLinkData[fKey]:
                            if x not in ts.dChanged: ts.dChanged[x] = ts.hdTitle[x]
                        break
                    
                    if len(tValue) > 0:
                        nList = list()
                        
                        for tKey in tList:
                            bValueZeroSkip = False
                            match ts.dDataType[x]:
                                case "int" | "float":
                                    if int(tKey) == 0: bValueZeroSkip = True
                            
                            if bValueZeroSkip == False:
                                if tKey in dLinkData:
                                    nList.append(dLinkData[tKey])
                                else:
                                    logs.Add(fullPath, fileName, orgSheetName, ts.hdTitle[x], orgY, orgX, "참조하는 값을 찾을 수 없습니다 (" + tKey + ")")
                            else:
                                if str(tKey) == str(0):
                                    nList.append(tKey)
                                else:
                                    if tKey in dLinkData:
                                        nList.append(dLinkData[tKey])
                                    else:
                                        logs.Add(fullPath, fileName, orgSheetName, ts.hdTitle[x], orgY, orgX, "참조하는 값을 찾을 수 없습니다 (" + tKey + ")")
                            
                        if x in ts.dChanged:
                            nValue = ListToString(nList, ",")
                            dbChangeValue[sheetName + "/" + str(y) + "/" + str(x)] = nValue
                    else:
                        if x in ts.dChanged:
                            if x in ts.dArray:
                                dbChangeValue[sheetName + "/" + str(y) + "/" + str(x)] = ""
                            else:
                                match ts.dDataType[x]:
                                    case "localtext":
                                        dbChangeValue[sheetName + "/" + str(y) + "/" + str(x)] = ""
                                    case _:
                                        dbChangeValue[sheetName + "/" + str(y) + "/" + str(x)] = 0
                    
                else:
                    logs.Add(fullPath, fileName, orgSheetName, ts.hdTitle[x], orgY, orgX, "참조하는 필드가 없습니다")
    
    return sheetName, ts, sd, logs, dbChangeValue


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Merge ArraySize (From Multiprocess result)
def MergeArraySize(dbArraySize: cArraySizeRefCheck):
    global p
    
    for tKey in dbArraySize.dArraySize:
        if tKey not in p.dbArraySize.dArraySize:
            p.dbArraySize.dArraySize[tKey] = dbArraySize.dArraySize[tKey]


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Merge Origin Data (From Multiprocess result)
def MergeOrigin(dbOrigin: cReferenceDataList):
    global p
    
    for tKey in dbOrigin.dOrigin:
        if tKey not in p.dbOrigin.dOrigin:
            p.dbOrigin.dOrigin[tKey] = dbOrigin.dOrigin[tKey]


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Merge Error Logs (from Multiprocess result)
def MergeLog(logs: cLogList):
    global p
    
    for l in logs.logs:
        p.listLog.Merge(logs.logs[l].fullPath, logs.logs[l].fileName, logs.logs[l].sheetName, logs.logs[l].colName, logs.logs[l].y, logs.logs[l].x, logs.logs[l].msg)


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Merge dbChangeValue (from Multiprocess result)
def MergeChangeValue(dbChangeValue: dict):
    global p
    
    for tKey in dbChangeValue:
        p.dbChangeValue[tKey] = dbChangeValue[tKey]


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Merge Table by mergeName
def MergeTables(pBar, mergeName: str):
    global p
    
    GenerateMergeTS(pBar, mergeName)
    if p.listLog.LogCount() == 0:
        GenerateMergeSD(pBar, mergeName)


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : TS Merge
def GenerateMergeTS(pBar, mergeName: str):
    global p
    
    maxX = 0
    
    pBar.strTitle = "Merge " + mergeName + " Table Structure"
    
    mergeTable = p.dbMerge[mergeName]
    
    pBar.t_count = len(mergeTable.dSheet)
    pBar.count = 0
    pBar.SetText(pBar.strTitle, "")
    pBar.SetRate(0)

    # Merge Table Structure
    for sheetName in mergeTable.dSheet:
        tempTS = mergeTable.dTS[sheetName]
        
        for tKey in tempTS.dCheckedX:
            mergeTable.dOldTitlePos[sheetName + "/" + tempTS.hdTitle[tKey]] = tKey
            
            if tempTS.hdTitle[tKey] not in mergeTable.dTitle:
                x = len(mergeTable.dTitle) + 1
                mergeTable.dTitle[tempTS.hdTitle[tKey]] = x
                
                # Copy Table Structure
                mergeTable.ts.hdTitle[x] = tempTS.hdTitle[tKey]
                mergeTable.ts.hdParam[x] = tempTS.hdParam[tKey]
                mergeTable.ts.hdExport[x] = tempTS.hdExport[tKey]
                mergeTable.ts.dCheckedX[x] = tempTS.dCheckedX[tKey]
                mergeTable.ts.dDataType[x] = tempTS.dDataType[tKey]
                if tKey in tempTS.dArray: mergeTable.ts.dArray[x] = tempTS.dArray[tKey]
                if tKey in tempTS.dOrigin: mergeTable.ts.dOrigin[x] = mergeName + "/" + tempTS.hdTitle[tKey]
                if tKey in tempTS.dUnique: mergeTable.ts.dUnique[x] = tempTS.dUnique[tKey]
                if tKey in tempTS.dEnum: mergeTable.ts.dEnum[x] = tempTS.dEnum[tKey]
                if tKey in tempTS.dEnumRef: mergeTable.ts.dEnumRef[x] = tempTS.dEnumRef[tKey]
                if tKey in tempTS.dValidCheck: mergeTable.ts.dValidCheck[x] = tempTS.dValidCheck[tKey]
                if tKey in tempTS.dPrefix: mergeTable.ts.dPrefix[x] = tempTS.dPrefix[tKey]
                if tKey in tempTS.dSuffix: mergeTable.ts.dSuffix[x] = tempTS.dSuffix[tKey]
                if tKey == tempTS.posKey: mergeTable.ts.posKey = x
                if maxX < x: maxX = x
                
                match(str(mergeTable.ts.hdExport[x]).lower()):
                    case "data": mergeTable.ts.cExportX.append(x)
                    case "check": mergeTable.ts.cCheckX.append(x)
        
        pBar.count += 1
        pBar.SetRate(pBar.count / pBar.t_count * 50)
    
    # Check Table Structures with Merged Table Structure
    pBar.count = 0
    for sheetName in mergeTable.dSheet:
        tempTS = mergeTable.dTS[sheetName]
        
        for tKey in tempTS.dCheckedX:
            x = mergeTable.dTitle[tempTS.hdTitle[tKey]]
            
            if tKey in tempTS.dChange:
                v = mergeTable.dTitle[tempTS.hdTitle[tempTS.dChange[tKey]]]
                if x not in mergeTable.ts.dChange: mergeTable.ts.dChange[x] = v
                if x not in mergeTable.ts.dChanged: mergeTable.ts.dChanged[x] = v
            
            if mergeTable.ts.hdTitle[x] != tempTS.hdTitle[tKey]:
                p.listLog.Add(tempTS.fullPath, tempTS.fileName, tempTS.sheetName, tempTS.hdTitle[tKey], 1, int(tKey), "필드 이름이 일치하지 않습니다")
            if mergeTable.ts.hdExport[x] != tempTS.hdExport[tKey]:
                p.listLog.Add(tempTS.fullPath, tempTS.fileName, tempTS.sheetName, tempTS.hdTitle[tKey], 3, int(tKey), "JSON 변환 여부가 일치하지 않습니다")
            if mergeTable.ts.dDataType[x] != tempTS.dDataType[tKey]:
                p.listLog.Add(tempTS.fullPath, tempTS.fileName, tempTS.sheetName, tempTS.hdTitle[tKey], 2, int(tKey), "데이터 타입이 일치하지 않습니다")
            
            # dEnum Check
            if x in mergeTable.ts.dEnum:
                if tKey not in tempTS.dEnum:
                    p.listLog.Add(tempTS.fullPath, tempTS.fileName, tempTS.sheetName, tempTS.hdTitle[tKey], 2, int(tKey), "ENUM 필드가 일치하지 않습니다")
                elif mergeTable.ts.dEnum[x] != tempTS.dEnum[tKey]:
                    p.listLog.Add(tempTS.fullPath, tempTS.fileName, tempTS.sheetName, tempTS.hdTitle[tKey], 2, int(tKey), "ENUM 필드가 일치하지 않습니다")

        pBar.count += 1
        pBar.SetRate(pBar.count / pBar.t_count * 50 + 50)
            
    mergeTable.ts.fileName = ""
    mergeTable.ts.fullPath = ""
    mergeTable.ts.sheetName = mergeName
    mergeTable.ts.maxCol = maxX
    
    p.dbMerge[mergeName] = mergeTable


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : SD Merge    
def GenerateMergeSD(pBar, mergeName: str):
    global p

    pBar.strTitle = "Merge " + mergeName + " Data"
    
    mergeTable = p.dbMerge[mergeName]
    
    pBar.t_count = len(mergeTable.dSheet)
    pBar.count = 0
    pBar.SetText(pBar.strTitle, f"0 / {pBar.t_count}")
    pBar.SetRate(0)
    
    for sheetName in mergeTable.dSheet:
        tempTS = mergeTable.dTS[sheetName]
        tempSD = mergeTable.dSD[sheetName]
        
        for orgY in tempSD.data:
            values = tempSD.data[orgY]
            y = len(mergeTable.sd.data) + 1
            
            dRow = dict()
            dOrgData = dict()
            
            for tTitle in mergeTable.dTitle:
                x = mergeTable.dTitle[tTitle]
                orgX = tempTS.GetPosByName(tTitle)
                
                if orgX > 0:
                    value = values[orgX]
                else:
                    match(str(mergeTable.ts.dDataType[x]).lower()):
                        case "int" | "float" | "change":
                            value = "0"
                        case "string" | "localtext" | "stream":
                            value = ""
                        case "bool":
                            value = "false"
                        case "type":
                            value = "NONE"
                        case _:
                            value = "0"
                
                dRow[x] = value
                
                if tTitle == mergeTable.mergeGroup and value not in mergeTable.dGroupID: mergeTable.dGroupID[value] = value
                
                tData = cMergeTableData()
                tData.SetData(tempTS.fullPath, tempTS.fileName, sheetName, tTitle, orgY, orgX, value)
                dOrgData[x] = tData
            
            mergeTable.sd.data[y] = dRow
            mergeTable.data[y] = dOrgData
            
            for tKey in mergeTable.ts.dOrigin:
                if tKey in mergeTable.ts.dChange:
                    p.dbOrigin.AddOrigin(mergeTable.ts.dOrigin[tKey], dRow[tKey], dRow[mergeTable.ts.dChange[tKey]])
                else:
                    p.dbOrigin.AddOrigin(mergeTable.ts.dOrigin[tKey], dRow[tKey], "")
        
        pBar.count += 1
        pBar.SetText(pBar.strTitle, f"{pBar.count} / {pBar.t_count}")
        pBar.SetRate(pBar.count / pBar.t_count * 100)
    
    mergeTable.ts.maxRow = len(mergeTable.sd.data)
    
    p.dbMerge[mergeName] = mergeTable


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Save Sheet Data to Memory
def SaveTableData(ts: cTableSchema, sd: cSheetData, tKey: str):
    global p
    
    tData = dict()
    if tKey in p.dbData:
        tData = p.dbData[tKey]
    
    tValue = dict()
    tValue["ts"] = ts
    tValue["sd"] = sd
    tData[ts.sheetName] = tValue
    p.dbData[tKey] = tData


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Data to Json Files
def MakeJsonFiles(pBar):
    global p
    
    pBar.strTitle = "Export Json"
    pBar.t_count = 0
    pBar.count = 0
    pBar.SetText(pBar.strTitle, "")
    pBar.SetRate(0)

    if p.bEnum == True: pBar.t_count += 2
    if p.bText == True: pBar.t_count += 1
    pBar.t_count += len(p.lMerge)
    if p.bData == True: pBar.t_count += len(p.dbData["data"])   
    
    if p.bEnum == True: MakeEnumCodeFile(pBar)
    if p.bText == True: MakeLocaltextJsonFile(pBar)
    if len(p.lMerge) > 0:
        for mergeName in p.lMerge:
            mergeTable = p.dbMerge[mergeName]
            if int(mergeTable.mergeCount) > 0:
                DataToSplitJson(mergeTable.ts, mergeTable.sd, mergeName, mergeTable.mergeGroup, mergeTable.mergeCount, pBar)
            else:
                DataToJson(mergeTable.ts, mergeTable.sd, mergeName)
    if p.bData == True: MakeDataJsonFile(pBar)


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Create Enum File
def MakeEnumCodeFile(pBar):


    pBar.SetText(pBar.strTitle, f"{pBar.count} / {pBar.t_count}")
    pBar.SetRate(pBar.count / pBar.t_count * 100)

    MakeClientEnumCodeFile()

    pBar.count += 1
    pBar.SetText(pBar.strTitle, f"{pBar.count} / {pBar.t_count}")
    pBar.SetRate(pBar.count / pBar.t_count * 100)

    #MakeServerEnumCodeFile()

    #pBar.count += 1
    #pBar.SetText(pBar.strTitle, f"{pBar.count} / {pBar.t_count}")
    #pBar.SetRate(pBar.count / pBar.t_count * 100)

# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Create Client Enum File
def MakeClientEnumCodeFile():
    global p

    if p.uiWin.strNamespace:
        strNameSpaceIndent = "\t"
    else:
        strNameSpaceIndent = ""

    str_list = list()
    if p.uiWin.strNamespace:
        str_list.append("namespace "+ p.uiWin.strNamespace +"\n{")
    bFirst = True

    for sType in p.dbEnum.dEnum:
        dValue = p.dbEnum.dEnum[sType]

        if p.dbEnum.dEnumNocode[sType] :
            continue

        if bFirst == True:
            sRow = strNameSpaceIndent + "public enum " + sType.upper() + "\n"+ strNameSpaceIndent + "{"
            bFirst = False
        else:
            sRow = "\n" + strNameSpaceIndent +"public enum " + sType.upper() + "\n" + strNameSpaceIndent +"{"
        
        for sText in dValue:
            sValue = str(dValue[sText])
            sComment = str(p.dbEnum.dComment[str(sType) + "/" + sValue])
            csText = str(sText).upper() + " = " + sValue + ","
            sRow = sRow + "\n"+ strNameSpaceIndent +"\t" + csText
            if len(sComment) > 0:
                sRow = sRow + GetTabStr(52, len(csText), 0) + "// " + sComment
        
        sRow = sRow + "\n" + strNameSpaceIndent + "}"
        if len(sRow) > 0:
            str_list.append("\n" + sRow)
    if p.uiWin.strNamespace:
        str_list.append("\n" + "}")

    if not bFirst:
        file = open(p.pathOutputScript + "TableEnum.cs", "w", encoding="utf-8")
        file.writelines(str_list)
        file.close()

# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Create Server Enum File
def MakeServerEnumCodeFile():
    global p
    
    str_list = list()
    bFirst = True
    
    for sType in p.dbEnum.dEnum:
        dValue = p.dbEnum.dEnum[sType]
        
        if bFirst == True:
            sRow = "export enum " + sType.upper() + " {"
            bFirst = False
        else:
            sRow = "\nexport enum " + sType.upper() + " {"
        
        for sText in dValue:
            sValue = str(dValue[sText])
            sComment = str(p.dbEnum.dComment[str(sType) + "/" + sValue])
            csText = str(sText).upper() + " = " + sValue + ","
            sRow = sRow + "\n\t" + csText
            if len(sComment) > 0:
                sRow = sRow + GetTabStr(56, len(csText), 0) + "// " + sComment
        
        sRow = sRow + "\n" + "}"
        if len(sRow) > 0:
            str_list.append("\n" + sRow)
    
    str_list.append("\n" + "}")
    
    file = open(p.pathOutput + "table_enum.ts", "w", encoding="utf-8")
    file.writelines(str_list)    
    file.close()


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Create Local Text File (Ordered by key string)
def MakeLocaltextJsonFile(pBar):
    global p
    
    if len(p.dbText.dKeyList) > 0:
        dKeyList = {key: value for key, value in sorted(p.dbText.dKeyList.items())}
        
        for tLang in p.dbText.dLocalText:
            str_list = list()
            bFirst = True
            
            str_list.append("[")
            
            for tKey in dKeyList:
                sRow = "\t\t\"key\": \"" + tKey + "\""
                sRow = sRow + ",\n\t\t\"text\": \"" + p.dbText.dLocalText[tLang][tKey] + "\""
                sRow = sRow + ",\n\t\t\"text_category\": \"" + p.dbText.dKeyList[tKey] + "\""
                if len(sRow) > 0:
                    if bFirst == True:
                        sRow = "\n\t{\n" + sRow + "\n\t}"
                        bFirst = False
                    else:
                        sRow = ",\n\t{\n" + sRow + "\n\t}"
                str_list.append(sRow)
            
            str_list.append("]")

            file = open(p.pathOutput + "string_" + tLang + ".json", "w", encoding="utf-8")
            file.writelines(str_list)    
            file.close()
    
    pBar.count += 1
    pBar.SetText(pBar.strTitle, f"{pBar.count} / {pBar.t_count}")
    pBar.SetRate(pBar.count / pBar.t_count * 100)


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Data Json File Generate
def MakeDataJsonFile(pBar):
    global p
    
    if "data" in p.dbData:
        tData = p.dbData["data"]
        
        for tKey in tData:
            ts = tData[tKey]["ts"]
            sd = tData[tKey]["sd"]
            DataToJson(ts, sd, ts.sheetName)
            if not ts.bNocode:
                SchemaToCSharpCode(ts, sd, ts.sheetName)
            pBar.count += 1
            pBar.SetText(pBar.strTitle, f"{pBar.count} / {pBar.t_count}")
            pBar.SetRate(pBar.count / pBar.t_count * 100)            


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Data To Json File
def DataToJson(ts: cTableSchema, sd: cSheetData, sheetName: str):
    global p
    
    if len(ts.cExportX) > 0:
        str_list = list()
        bFirst = True

        if not ts.bValueOnly:
            str_list.append("[")
        
        for y in sd.data:
            bFirstLine = True
            sRow = ""
            dRow = sd.data[y]
            if not ts.bValueOnly :
                tValue = dRow[ts.posKey]
            else :
                tValue = ""

            if len(tValue) > 0 or ts.bValueOnly :
                for x in ts.cExportX:
                    tValue = dRow[x]
                    if str(tValue).lower() != "null" :
                        if x in ts.dChanged:
                            if sheetName + "/" + str(y) + "/" + str(x) in p.dbChangeValue:
                                tValue = p.dbChangeValue[sheetName + "/" + str(y) + "/" + str(x)]
                            if x in ts.dReplaceComma:
                                sRow = sRow + GetJsonLine(ts.hdTitle[x], str(tValue), "int", x in ts.dArray, bFirstLine, True, ts.dReplaceComma[x])
                            else:
                                sRow = sRow + GetJsonLine(ts.hdTitle[x], str(tValue), "int", x in ts.dArray, bFirstLine, False, "")
    
                            tValue = dRow[x]
                            # if x in ts.dReplaceComma:
                            #     sRow = sRow + GetJsonLine(ts.hdTitle[x] + "_desc", str(tValue), "string", x in ts.dArray, False, True, ts.dReplaceComma[x])
                            # else:
                            #     sRow = sRow + GetJsonLine(ts.hdTitle[x] + "_desc", str(tValue), "string", x in ts.dArray, False, False, "")

                        else:
                            if x in ts.dReplaceComma:
                                sRow = sRow + GetJsonLine(ts.hdTitle[x], str(tValue), ts.dDataType[x], x in ts.dArray, bFirstLine, True, ts.dReplaceComma[x])
                            else:
                                sRow = sRow + GetJsonLine(ts.hdTitle[x], str(tValue), ts.dDataType[x], x in ts.dArray, bFirstLine, False, "")
                        
                        bFirstLine = False
                
            if len(sRow) > 0:
                if bFirst == True:
                    str_list.append("\n\t{")
                    bFirst = False
                else:
                    str_list.append(",\n\t{")
                
                str_list.append("\n"+sRow+"\n\t}")

        if not ts.bValueOnly:
            str_list.append("\n]")

        if ts.bValueOnly:
            text = "".join(str_list)
            lines = text.splitlines()
            result = '\n'.join([line[1:] if line.startswith('\t') else line for line in lines])
            str_list = result[1:] if result.startswith('\n') else result

        finalFolder =  p.pathOutputJson
        if ts.subFolder != "" :
            finalFolder = finalFolder + ts.subFolder
            os.makedirs(os.path.dirname(finalFolder), exist_ok = True)

        file = open(finalFolder + snake_to_pascal(sheetName) + "Table" + ".json", "w", encoding="utf-8")
        file.writelines(str_list)
        file.close()


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Data To Json File
def SchemaToCSharpCode(ts: cTableSchema, sd: cSheetData, sheetName: str):
    global p

    if len(ts.cExportX) > 0:
        str_list = list()

        if p.uiWin.strNamespace:
            strNameSpaceIndent = "\t"
        else:
            strNameSpaceIndent = ""

        str_list.append("using System;\n")
        # str_list.append("using System.Collections.Generic\n{")
        str_list.append("using UnityEngine;\n")
        str_list.append("using DUG;\n")
        str_list.append("\n")
        if p.uiWin.strNamespace:
            str_list.append("namespace " + p.uiWin.strNamespace + "\n{\n")

        str_list.append(strNameSpaceIndent +"[PreferBinarySerialization]\n")

        tableClassName = snake_to_pascal(sheetName) + "Table"
        recordClassName = tableClassName + "Record"

        if not ts.bValueOnly:
            keyFieldName = ts.hdTitle[ts.posKey]
            keyFieldType = ts.dDataType[ts.posKey]
            if keyFieldType == "type" :
                keyFieldType = ts.dEnum[ts.posKey]

        if ts.bValueOnly :
            strTableClass = strNameSpaceIndent + "public partial class " + tableClassName + " : ValueOnlyTable<" + recordClassName + ">\n"
        else :
            strTableClass = strNameSpaceIndent + "public partial class " + tableClassName + " : KeyValueTable<" + keyFieldType + ", " + recordClassName +">\n"

        strTableClass += strNameSpaceIndent + "{\n"

        if not ts.bValueOnly :
            strTableClass += strNameSpaceIndent + "\tpublic " + tableClassName +"()" + " : " + "base(nameof("+ recordClassName +"."+ keyFieldName + "))\n"
            strTableClass += strNameSpaceIndent +"\t{}\n"

        strTableClass += strNameSpaceIndent +"}\n"

        str_list.append(strTableClass)
        str_list.append("\n")

        str_list.append(strNameSpaceIndent + "[Serializable]\n")

        strRecordClass = strNameSpaceIndent +"public partial class " + recordClassName + "\n"
        strRecordClass += strNameSpaceIndent + "{" + "\n"
        for x in ts.cExportX:

             if ts.dDataType[x] == "type" :
                 dataTypeText = ts.dEnum[x]
             else :
                 dataTypeText = ts.dDataType[x]

             if x in ts.dArray :
                 csText = dataTypeText+ "[] " + ts.hdTitle[x] + ";"
             else :
                 csText = dataTypeText + " " + ts.hdTitle[x] + ";"

             strRecordClass += strNameSpaceIndent + "\tpublic " + csText + "\n"
        strRecordClass += strNameSpaceIndent + "}\n"
        str_list.append(strRecordClass)

        if p.uiWin.strNamespace:
            str_list.append("}\n")

        finalFolder =  p.pathOutputScript
        if ts.subFolder != "" :
            finalFolder = finalFolder + ts.subFolder
            os.makedirs(os.path.dirname(finalFolder), exist_ok = True)

        file = open(finalFolder + tableClassName + ".cs", "w", encoding="utf-8")
        file.writelines(str_list)
        file.close()

# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Data To Merge Table Split Json File
def DataToSplitJson(ts: cTableSchema, sd: cSheetData, sheetName: str, indexKeyName: str, mergeCount: int, pBar):
    global p
    
    if len(ts.cExportX) > 0:
        posIndexKey = ts.GetPosByName(indexKeyName)
        idx = cIndex()
        idx.Init(int(mergeCount))

        dKeyList = dict()
        for y in sd.data:
            if int(sd.data[y][posIndexKey]) in dKeyList:
                dKeyList[int(sd.data[y][posIndexKey])].append(y)
            else:
                nList = list()
                nList.append(y)
                dKeyList[int(sd.data[y][posIndexKey])] = nList
        
        sKeyList = {key: value for key, value in sorted(dKeyList.items())}
        
        for tKey in sKeyList:
            for y in sKeyList[tKey]:
                tIndexKey = sd.data[y][posIndexKey]
                idx.AddIndexKey(int(tIndexKey), y)
        
        for iKey in idx.idxDataLine:
            yList = idx.idxDataLine[iKey]
            fName = sheetName + "_" + "{:05d}".format(iKey)
            
            str_list = list()
            str_list.append("[")
            bFirst = True
            
            for y in yList:
                bFirstLine = True
                sRow = ""
                dRow = sd.data[y]
                if len(dRow[ts.posKey]) > 0:
                    for x in ts.cExportX:
                        tValue = dRow[x]
                        if str(tValue).lower() != "null" :
                            if x in ts.dChanged:
                                if sheetName + "/" + str(y) + "/" + str(x) in p.dbChangeValue:
                                    tValue = p.dbChangeValue[sheetName + "/" + str(y) + "/" + str(x)]
                                if x in ts.dReplaceComma:
                                    sRow = sRow + GetJsonLine(ts.hdTitle[x], str(tValue), "int", x in ts.dArray, bFirstLine, True, ts.dReplaceComma[x])
                                else:
                                    sRow = sRow + GetJsonLine(ts.hdTitle[x], str(tValue), "int", x in ts.dArray, bFirstLine, False, "")
        
                                tValue = dRow[x]
                                # if x in ts.dReplaceComma:
                                #     sRow = sRow + GetJsonLine(ts.hdTitle[x] + "_desc", str(tValue), "string", x in ts.dArray, False, True, ts.dReplaceComma[x])
                                # else:
                                #     sRow = sRow + GetJsonLine(ts.hdTitle[x] + "_desc", str(tValue), "string", x in ts.dArray, False, False, "")

                            else:
                                if x in ts.dReplaceComma:
                                    sRow = sRow + GetJsonLine(ts.hdTitle[x], str(tValue), ts.dDataType[x], x in ts.dArray, bFirstLine, True, ts.dReplaceComma[x])
                                else:
                                    sRow = sRow + GetJsonLine(ts.hdTitle[x], str(tValue), ts.dDataType[x], x in ts.dArray, bFirstLine, False, "")
                            
                            bFirstLine = False
                    
                    if len(sRow) > 0:
                        if bFirst == True:
                            str_list.append("\n\t{")
                            bFirst = False
                        else:
                            str_list.append(",\n\t{")
                        
                        str_list.append("\n"+sRow+"\n\t}")

            str_list.append("\n]")
                
            file = open(p.pathOutput + fName + ".json", "w", encoding="utf-8")
            file.writelines(str_list)
            file.close()                        
        
        MakeSplitJsonIndex(sheetName, idx, indexKeyName)
    
    pBar.count += 1
    pBar.SetText(pBar.strTitle, f"{pBar.count} / {pBar.t_count}")
    pBar.SetRate(pBar.count / pBar.t_count * 100)

# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Index Navigation Json
def MakeSplitJsonIndex(sheetName: str, idx: cIndex, keyName: str):
    global p
    
    str_list = list()
    bFirst = True
    str_list.append("[")    
    
    for tKey in idx.idx:
        sRow = ""
        
        if len(str(tKey)) > 0:
            sRow = sRow + GetJsonLine(keyName, str(tKey), "int", False, True, False, "")            
            fName = sheetName + "_" + "{:05d}".format(idx.idx[tKey]) + ".json"
            sRow = sRow + GetJsonLine("filename", fName, "string", False, False, False, "")
        
        if len(sRow) > 0:
            if bFirst == True:
                str_list.append("\n\t{")
                bFirst = False
            else:
                str_list.append(",\n\t")
            str_list.append("\n"+sRow+"\n\t}")
    
    str_list.append("\n]")
    
    file = open(p.pathOutput + sheetName + "_index.json", "w", encoding="utf-8")
    file.writelines(str_list)
    file.close()


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Make Data to Json Text Line
def GetJsonLine(sTitle: str, sValue: str, sDataType: str, bArray: bool, bFirstLine: bool, bReplaceComma: bool, sReplaceComma: str):
    res = ""
    
    match sDataType.lower():
        case "string" | "localtext":
            sValue = "\"" + sValue + "\""
    
    if bFirstLine == False: res = ",\n"
    
    if bArray == True:
        if len(sValue) > 0 and sValue != "\"\"":
            res = res + "\t\t\"" + sTitle + "\": ["
        
            match sDataType.lower():
                case "string" | "localtext":
                    if bReplaceComma == True:
                        res = res + "\n\t\t\t" + sValue.replace(",", "\",\n\t\t\t\"").replace(sReplaceComma, ",")
                    else:
                        res = res + "\n\t\t\t" + sValue.replace(",", "\",\n\t\t\t\"")
                case _:
                    if bReplaceComma == True:
                        res = res + "\n\t\t\t" + sValue.replace(",", ",\n\t\t\t").replace(sReplaceComma, ",")
                    else:
                        res = res + "\n\t\t\t" + sValue.replace(",", ",\n\t\t\t")
        
            res = res + "\n\t\t]"
        else:
            res = res + "\t\t\"" + sTitle + "\": []"
    else:
        if bReplaceComma == True:
            res = res + "\t\t\"" + sTitle + "\": " + sValue.replace(sReplaceComma, ",")
        else:
            res = res + "\t\t\"" + sTitle + "\": " + sValue
    
    return res


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : Error logs to Excel file
def SaveErrorLogFile(pBar):
    global p
    
    pBar.strTitle = "Write Error Logs"
    pBar.t_count = len(p.listLog.logs)
    pBar.count = 0
    pBar.SetText(pBar.strTitle, f"{pBar.count} / {pBar.t_count}")
    pBar.SetRate(0)
    lenCell = dict()
    if pBar.t_count > 0:
        wb = Workbook()
        ws = wb.active
        
        ws.cell(row=1, column=1).value = "File Name"
        ws.cell(row=1, column=2).value = "Sheet Name"
        ws.cell(row=1, column=3).value = "Data Name"
        ws.cell(row=1, column=4).value = "Column"
        ws.cell(row=1, column=5).value = "Row"
        ws.cell(row=1, column=6).value = "Log"
        ws.cell(row=1, column=7).value = "Link"
        
        hdFont = Font(bold=True)
        hdColor = PatternFill(patternType="solid", fgColor="E2EFDA")
        hdAlign = Alignment(horizontal="center", vertical="center")
        
        for x in range(1, 8):
            ws.cell(row=1, column=x).font = hdFont
            ws.cell(row=1, column=x).fill = hdColor
            ws.cell(row=1, column=x).alignment = hdAlign
            lenCell[x] = len(str(ws.cell(row=1, column=x).value))
        
        for idx in p.listLog.logs:
            log = p.listLog.logs[idx]
            ws.cell(row=pBar.count + 2, column=1).value = log.fileName
            ws.cell(row=pBar.count + 2, column=2).value = log.sheetName
            ws.cell(row=pBar.count + 2, column=3).value = log.colName
            ws.cell(row=pBar.count + 2, column=4).value = log.x
            ws.cell(row=pBar.count + 2, column=5).value = log.y
            ws.cell(row=pBar.count + 2, column=6).value = log.msg
            cellAddress = Xlref(log.y, log.x, False)
            ws.cell(row=pBar.count + 2, column=7).value = f'=HYPERLINK("[{log.fullPath}/{log.fileName}]{log.sheetName}!{cellAddress}", "LINK")'
            ws.cell(row=pBar.count + 2, column=7).style = "Hyperlink"
            
            for x in range(1, 7):
                if lenCell[x] < len(str(ws.cell(row=pBar.count + 2, column=x).value)):
                    lenCell[x] = len(str(ws.cell(row=pBar.count + 2, column=x).value))
            
            pBar.count += 1
            pBar.SetText(pBar.strTitle, f"{pBar.count} / {pBar.t_count}")
            pBar.SetRate(pBar.count / pBar.t_count * 100)
        
        ws.column_dimensions['A'].width = lenCell[1] * 1.3
        ws.column_dimensions['B'].width = lenCell[2] * 1.3
        ws.column_dimensions['C'].width = lenCell[3] * 1.3
        ws.column_dimensions['D'].width = lenCell[4] * 1.3
        ws.column_dimensions['E'].width = lenCell[5] * 1.3
        ws.column_dimensions['F'].width = lenCell[6] * 1.3
        ws.column_dimensions['G'].width = lenCell[7] * 1.3
        
        p.errorFileName = "ExportError_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
        fileName = p.pathError + p.errorFileName
        wb.save(fileName)
        wb.close()
        
        #subprocess.check_call(["open", "-a", "Microfost Excel", fileName])
        #os.system(fileName)
        subprocess.Popen(['start',fileName],stdout=subprocess.PIPE,shell = True)
    

################################################################################################################
# UI CLASS & FUNCTION
# --------------------------------------------------------------------------------------------------------------
# UI FUNCTION : Move window position to screen center
def Center(uiWin):
    uiWin.update_idletasks()
    width = uiWin.winfo_width()
    frm_width = uiWin.winfo_rootx() - uiWin.winfo_x()
    win_width = width + 2 * frm_width
    height = uiWin.winfo_height()
    titlebar_height = uiWin.winfo_rooty() - uiWin.winfo_y()
    win_height = height + titlebar_height + frm_width
    x = uiWin.winfo_screenwidth() // 2 - win_width // 2
    y = uiWin.winfo_screenheight() // 2 - win_height // 2
    uiWin.geometry('{}x{}+{}+{}'.format(width, height, x, y))
    uiWin.deiconify()


# --------------------------------------------------------------------------------------------------------------
# UI CLASS : Progress popup
class cProgressPopup(object):
    def __init__(self, master):
        self.state = ""
        self.pRate = DoubleVar()
        
        self.strTitle = ""
        self.t_count = 0
        self.count = 0
        
        top = self.top = Toplevel(master.main)
        top.geometry("600x115")
        top.wm_overrideredirect(True)
        top.config(bg="#E0E0E0")

        top.lift()
        top.focus_force()
        top.grab_set()
                
        self.f1 = Frame(top, bd=0, bg="#E0E0E0")
        self.f2 = Frame(top, bd=0, bg="#E0E0E0")
        
        self.f1.grid(row=0, column=0, padx=8, pady=4, sticky=NSEW)
        self.f2.grid(row=1, column=0, padx=8, pady=4, sticky=NSEW)
        
        self.title = Label(self.f1, text="", bg="#E0E0E0", anchor=CENTER, font='sans 10 bold')
        self.textDesc = Label(self.f1, text="", bg="#E0E0E0", anchor=CENTER)
        self.textRate = Label(self.f1, text="", bg="#E0E0E0", anchor=CENTER)
        
        self.title.grid(row=0, column=0, padx=4, pady=4, sticky=NSEW)
        self.textDesc.grid(row=1, column=0, padx=4, pady=0, sticky=NSEW)
        self.textRate.grid(row=2, column=0, padx=4, pady=0, sticky=NSEW)
        
        self.f1.rowconfigure((0,1,2), weight=1)
        self.f1.columnconfigure(0, weight=1)
        
        self.pBar = Progressbar(self.f2, maximum=100, variable=self.pRate)
        self.pBar.grid(row=0, column=0, padx=4, pady=4, sticky=NSEW)

        self.f2.rowconfigure((0,1), weight=1)
        self.f2.columnconfigure(0, weight=1)
        
        top.rowconfigure(0, weight=1)
        top.columnconfigure(0, weight=1)
        
        Center(top)

    # CLASS FUNCTION : Init popup
    def SetInit(self, strTitle: str, strDesc: str, strRate: str, r: float):
        self.title["text"] = strTitle
        self.textDesc["text"] = strDesc
        self.textRate["text"] = strRate
        self.pRate.set(r)
        self.top.update()

    # CLASS FUNCTION : Change text
    def SetText(self, strDesc: str, strRate: str):
        self.textDesc["text"] = strDesc
        self.textRate["text"] = strRate
        self.top.update()

    # CLASS FUNCTION : Change bar rate
    def SetRate(self, rateBar: float):
        self.pRate.set(rateBar)
        self.top.update()
    
    # CLASS FUNCTION : Exit popup
    def Exit(self):
        self.top.destroy()


# --------------------------------------------------------------------------------------------------------------
# UI CLASS : Preset save popup
class cPresetSavePopup(object):
    def __init__(self, master):
        self.presetName = ""
        self.state = PresetSavePopupState.NORMAL
        
        top = self.top = Toplevel(master.main)
        top.geometry("500x500")
        top.wm_overrideredirect(True)
        top.config(bg="#E0E0E0")        

        top.lift()
        top.focus_force()
        top.grab_set()
        
        self.f1 = Frame(top, bd=0, bg="#E0E0E0")
        self.f2 = Frame(top, bd=0, bg="#E0E0E0")
        self.f3 = Frame(top, bd=0, bg="#E0E0E0")
        self.f4 = Frame(top, bd=0, bg="#E0E0E0")
        
        self.f1.grid(row=0, column=0, padx=8, pady=4, sticky=NSEW)
        self.f2.grid(row=1, column=0, padx=4, pady=4, sticky=NSEW)
        self.f3.grid(row=2, column=0, padx=8, pady=4, sticky=NSEW)
        self.f4.grid(row=3, column=0, padx=4, pady=4, sticky=NSEW)        
        
        # Input preset name
        self.inputLabel = Label(self.f1, text="INPUT PRESET NAME", bg="#E0E0E0", font='sans 10 bold')
        self.inputLabel.pack(padx=4, pady=4)
        
        self.inputField = Entry(self.f2)
        self.inputField.pack(padx=4, pady=4, fill=BOTH)

        # Checked file list
        self.scrollFile = Scrollbar(self.f3)
        self.scrollFile.pack(side=RIGHT, fill=Y)
        
        self.listFile = Listbox(self.f3, yscrollcommand=self.scrollFile.set)
        self.listFile.pack(side=LEFT, fill=BOTH, expand=YES)
        self.scrollFile.config(command=self.listFile.yview)
        
        # Bottom buttons
        self.btnSave = Button(self.f4, text="SAVE", width=180, height=1, command=self.Save)
        self.btnExit = Button(self.f4, text="CANCEL", width=180, height=1, command=self.Exit)
        
        self.btnSave.grid(row=0, column=0, padx=4, pady=4, sticky=NSEW)
        self.btnExit.grid(row=0, column=1, padx=4, pady=4, sticky=NSEW)
        
        self.f4.rowconfigure(0, weight=1)
        self.f4.columnconfigure((0,1), weight=1)
        
        top.rowconfigure(2, weight=1)
        top.columnconfigure((0,1,2,3), weight=1)
        
        Center(top)
        
        self.ShowFileList(master)

    # CLASS FUNCTION : Show checked file list when popup oepned
    def ShowFileList(self, master):
        if len(master.dCheckedFileList):
            for fileName in master.dCheckedFileList:
                self.listFile.insert(END, fileName)
        else:
            self.presetName = ""
            self.state = PresetSavePopupState.NO_FILE
            self.top.destroy()
    
    # CLASS FUNCTION : Click save butten event
    def Save(self):
        self.presetName = self.inputField.get()
        
        if len(self.presetName) > 0:
            self.state = PresetSavePopupState.SAVE
        else:
            self.state = PresetSavePopupState.NO_NAME
            
        self.top.destroy()
    
    # CLASS FUNCTION : Click exit butten event
    def Exit(self):
        self.presetName = ""
        self.state = PresetSavePopupState.EXIT
        self.top.destroy()


# --------------------------------------------------------------------------------------------------------------
# UI CLASS : Preset load popup
class cPresetLoadPopup(object):
    def __init__(self, master):
        self.presetName = ""
        self.stae = PresetLoadPopupState.NORMAL
        
        top = self.top = Toplevel(master.main)
        top.geometry("500x500")
        top.wm_overrideredirect(True)
        top.config(bg="#E0E0E0")

        top.lift()
        top.focus_force()
        top.grab_set()
                
        self.f1 = Frame(top, bd=0, bg="#E0E0E0")
        self.f2 = Frame(top, bd=0, bg="#E0E0E0")
        self.f3 = Frame(top, bd=0, bg="#E0E0E0")
        
        self.f1.grid(row=0, column=0, padx=8, pady=4, sticky=NSEW)
        self.f2.grid(row=1, column=0, padx=8, pady=4, sticky=NSEW)
        self.f3.grid(row=2, column=0, padx=4, pady=4, sticky=NSEW)
        
        # Preset List
        self.scrollPreset = Scrollbar(self.f1)
        self.scrollPreset.pack(side=RIGHT, fill=Y)
        
        self.listPreset = Treeview(self.f1, yscrollcommand=self.scrollPreset.set, selectmode="extended")
        self.listPreset.pack(side=LEFT, fill=BOTH, expand=YES)
        self.scrollPreset.config(command=self.listPreset.yview)
        
        self.listPreset["columns"] = ("Name")
        self.listPreset.column("#0", width=0, anchor=CENTER, stretch=NO)
        self.listPreset.column("Name", width=300, anchor=CENTER)
        
        self.listPreset.heading("#0", text="")
        self.listPreset.heading("Name", text="SELECT PRESET NAME", anchor=CENTER)
        self.listPreset.bind("<<TreeviewSelect>>", lambda e: self.ShowFileList(master, e))
        
        # Selected preset file list
        self.scrollFile = Scrollbar(self.f2)
        self.scrollFile.pack(side=RIGHT, fill=Y)
        
        self.listFile = Listbox(self.f2, yscrollcommand=self.scrollFile.set)
        self.listFile.pack(side=LEFT, fill=BOTH, expand=YES)
        self.scrollFile.config(command=self.listFile.yview)
        
        # Bottom buttons
        self.btnLoad = Button(self.f3, text="Load", width=180, height=1, command=self.Load)
        self.btnDelete = Button(self.f3, text="Delete", width=180, height=1, command=self.Delete)
        self.btnExit = Button(self.f3, text="Cancel", width=180, height=1, command=self.Exit)
        
        self.btnLoad.grid(row=0, column=0, padx=4, pady=4, sticky=NSEW)
        self.btnDelete.grid(row=0, column=1, padx=4, pady=4, sticky=NSEW)
        self.btnExit.grid(row=0, column=2, padx=4, pady=4, sticky=NSEW)
        
        self.f3.rowconfigure(0, weight=1)
        self.f3.columnconfigure((0,1,2), weight=1)
        
        top.rowconfigure(1, weight=1)
        top.columnconfigure((0,1), weight=1)

        Center(top)
                
        self.ShowPresetName(master)
    
    # CLASS Function : Show preset list when popup oepned
    def ShowPresetName(self, master):
        count = 0
        for fileName in master.dPreset:
            count = count + 1
            self.listPreset.insert(parent="", index="end", iid=count, values=(fileName))

    # CLASS Function : Show file list when preset name clicked
    def ShowFileList(self, master, event):
        sel = self.listPreset.focus()
        self.presetName = self.listPreset.item(sel)["values"][0]
        self.listFile.delete(0,END)
        for fileName in master.dPreset[self.presetName]:
            self.listFile.insert(END, fileName)
    
    # CLASS FUNCTION : Click load butten event
    def Load(self):
        sel = self.listPreset.focus()
        
        if len(sel) > 0:
            self.presetName = self.listPreset.item(sel)["values"][0]
            self.state = PresetLoadPopupState.LOAD
        else:
            self.presetName = ""
            self.state = PresetLoadPopupState.NO_NAME
        
        self.top.destroy()

    # CLASS FUNCTION : Click delete butten event
    def Delete(self):
        sel = self.listPreset.focus()
        
        if len(sel) > 0:
            self.presetName = self.listPreset.item(sel)["values"][0]
            self.state = PresetLoadPopupState.DELETE            
        else:
            self.presetName = ""
            self.state = PresetLoadPopupState.NO_NAME
        
        self.top.destroy()
    
    # CLASS FUNCTION : Click exit butten event
    def Exit(self):
        self.presetName = ""
        self.state = PresetLoadPopupState.EXIT
        self.top.destroy()


# --------------------------------------------------------------------------------------------------------------
# UI CLASS : Edit popup
class cEditPopup(object):
    def __init__(self, master):
        self.state = ""
        self.dExcelList = copy.deepcopy(master.dExcelList)
        self.dCheckedFileList = master.dCheckedFileList
        self.dListFlag = dict()
        
        self.fileName = ""
        
        top = self.top = Toplevel(master.main)
        top.geometry("800x500")
        top.wm_overrideredirect(True)
        top.config(bg="#E0E0E0")

        top.lift()
        top.focus_force()
        top.grab_set()
                
        self.f1 = Frame(top, bd=0, bg="#E0E0E0")
        self.f2 = Frame(top, bd=0, bg="#E0E0E0")
        self.f3 = Frame(top, bd=0, bg="#E0E0E0")
        
        self.f1.grid(row=0, column=0, padx=8, pady=8, sticky=NSEW)
        self.f2.grid(row=1, column=0, padx=8, pady=8, sticky=NSEW)
        self.f3.grid(row=2, column=0, padx=4, pady=4, sticky=NSEW)
        
        # Checked File List
        self.scrollFile = Scrollbar(self.f1)
        self.scrollFile.pack(side=RIGHT, fill=Y)
        
        self.listFile = Treeview(self.f1, yscrollcommand=self.scrollFile.set, selectmode="extended")
        self.listFile.pack(side=LEFT, fill=BOTH, expand=YES)
        self.scrollFile.config(command=self.listFile.yview)
        
        self.listFile["columns"] = ("Name", "Edited")
        self.listFile.column("#0", width=0, anchor=CENTER, stretch=NO)
        self.listFile.column("Name", width=300, anchor=CENTER)
        self.listFile.column("Edited", width=100, anchor=CENTER)
        
        self.listFile.heading("#0", text="")
        self.listFile.heading("Name", text="SELECT FILE NAME", anchor=CENTER)
        self.listFile.heading("Edited", text="EDIT FLAG", anchor=CENTER)
        self.listFile.bind("<<TreeviewSelect>>", self.SetExcelInfo)
        
        # Selected excel info
        self.l1 = Label(self.f2, text="Index", width=10, anchor=E, bg="#E0E0E0")
        self.l2 = Label(self.f2, text="Data type", width=10, anchor=E, bg="#E0E0E0")
        self.l3 = Label(self.f2, text="Merge Name", width=10, anchor=E, bg="#E0E0E0")
        self.l4 = Label(self.f2, text="Merge Group", width=10, anchor=E, bg="#E0E0E0")
        self.l5 = Label(self.f2, text="Merge Count", width=10, anchor=E, bg="#E0E0E0")
        self.l6 = Label(self.f2, text="Tag", width=10, anchor=E, bg="#E0E0E0")
        self.l7 = Label(self.f2, text="Sub Dir", width=10, anchor=E, bg="#E0E0E0")
        self.l8 = Label(self.f2, text="File Name", width=10, anchor=E, bg="#E0E0E0")
        self.l9 = Label(self.f2, text="Sheet List", width=10, anchor=E, bg="#E0E0E0")
        self.l10 = Label(self.f2, text="Desc", width=10, anchor=E, bg="#E0E0E0")
        
        self.i1 = Label(self.f2, text="", width=300, anchor=W, relief=SUNKEN)
        self.i2 = Label(self.f2, text="", width=300, anchor=W, relief=SUNKEN)
        self.i3 = Label(self.f2, text="", width=300, anchor=W, relief=SUNKEN)
        self.i4 = Label(self.f2, text="", width=300, anchor=W, relief=SUNKEN)
        self.i5 = Label(self.f2, text="", width=300, anchor=W, relief=SUNKEN)
        self.i6 = Entry(self.f2)
        self.i7 = Label(self.f2, text="", width=300, anchor=W, relief=SUNKEN)
        self.i8 = Label(self.f2, text="", width=300, anchor=W, relief=SUNKEN)
        self.i9 = Label(self.f2, text="", width=300, anchor=W, relief=SUNKEN)
        self.i10 = Entry(self.f2)
        
        self.l1.grid(row=0, column=0, padx=4, pady=4, sticky=NSEW)
        self.l2.grid(row=1, column=0, padx=4, pady=4, sticky=NSEW)
        self.l3.grid(row=2, column=0, padx=4, pady=4, sticky=NSEW)
        self.l4.grid(row=3, column=0, padx=4, pady=4, sticky=NSEW)
        self.l5.grid(row=4, column=0, padx=4, pady=4, sticky=NSEW)
        self.l6.grid(row=5, column=0, padx=4, pady=4, sticky=NSEW)
        self.l7.grid(row=6, column=0, padx=4, pady=4, sticky=NSEW)
        self.l8.grid(row=7, column=0, padx=4, pady=4, sticky=NSEW)
        self.l9.grid(row=8, column=0, padx=4, pady=4, sticky=NSEW)
        self.l10.grid(row=9, column=0, padx=4, pady=4, sticky=NSEW)
        
        self.i1.grid(row=0, column=1, padx=4, pady=4, sticky=NSEW)
        self.i2.grid(row=1, column=1, padx=4, pady=4, sticky=NSEW)
        self.i3.grid(row=2, column=1, padx=4, pady=4, sticky=NSEW)
        self.i4.grid(row=3, column=1, padx=4, pady=4, sticky=NSEW)
        self.i5.grid(row=4, column=1, padx=4, pady=4, sticky=NSEW)
        self.i6.grid(row=5, column=1, padx=4, pady=4, sticky=NSEW)
        self.i7.grid(row=6, column=1, padx=4, pady=4, sticky=NSEW)
        self.i8.grid(row=7, column=1, padx=4, pady=4, sticky=NSEW)
        self.i9.grid(row=8, column=1, padx=4, pady=4, sticky=NSEW)
        self.i10.grid(row=9, column=1, padx=4, pady=4, sticky=NSEW)
        
        self.f2.rowconfigure((0,1,2,3,4,5,6,7,8,9), weight=1)
        self.f2.columnconfigure(1, weight=1)
        
        # Bottom buttons
        self.btnSave = Button(self.f3, text="SAVE", width=180, height=1, command=self.Save)
        self.btnExit = Button(self.f3, text="CANCEL", width=180, height=1, command=self.Exit)
        
        self.btnSave.grid(row=0, column=0, padx=4, pady=4, sticky=NSEW)
        self.btnExit.grid(row=0, column=1, padx=4, pady=4, sticky=NSEW)
        
        self.f3.rowconfigure(0, weight=1)
        self.f3.columnconfigure((0,1), weight=1)
        
        top.rowconfigure((0,1), weight=1)
        top.columnconfigure(0, weight=1)
        
        self.ShowFileList()
        Center(top)

    # CLASS FUNCTION : Show checked file list when popup opened
    def ShowFileList(self):
        if len(self.dCheckedFileList) > 0:
            count = 0
            for fileName in self.dCheckedFileList:
                count = count + 1
                self.listFile.insert(parent="", index="end", iid=count, values=(fileName, ""))
            
            self.SetExcelInfoInit()
        else:
            self.state = EditPopupState.NO_FILE
            self.top.destroy()
    
    # CLASS FUNCTION : Set selected file info when popup opened
    def SetExcelInfoInit(self):
        self.listFile.selection_set(1)
        self.fileName = self.listFile.item("1")["values"][0]
        dExcelInfo = self.dExcelList[self.fileName]

        self.i1["text"] = dExcelInfo["index"]
        self.i2["text"] = dExcelInfo["type"]
        #self.i3["text"] = dExcelInfo["mergeName"]
        #self.i4["text"] = dExcelInfo["mergeGroup"]
        #self.i5["text"] = dExcelInfo["mergeCount"]
        self.i6.delete(0, "end")
        self.i6.insert(0, dExcelInfo["tag"])
        self.i7["text"] = dExcelInfo["subDir"]
        self.i8["text"] = dExcelInfo["fileName"]
        self.i9["text"] = dExcelInfo["sheetList"]
        self.i10.delete(0, "end")
        self.i10.insert(0, dExcelInfo["desc"])        
    
    # CLASS FUNCTION : Set selected file info
    def SetExcelInfo(self, event):
        self.SaveExcelInfo()
        
        sel = self.listFile.focus()
        if len(sel) > 0:
            self.fileName = self.listFile.item(sel)["values"][0]
            dExcelInfo = self.dExcelList[self.fileName]
            
            self.i1["text"] = dExcelInfo["index"]
            self.i2["text"] = dExcelInfo["type"]
            self.i3["text"] = dExcelInfo["mergeName"]
            self.i4["text"] = dExcelInfo["mergeGroup"]
            self.i5["text"] = dExcelInfo["mergeCount"]
            self.i6.delete(0, "end")
            self.i6.insert(0, dExcelInfo["tag"])
            self.i7["text"] = dExcelInfo["subDir"]
            self.i8["text"] = dExcelInfo["fileName"]
            self.i9["text"] = dExcelInfo["sheetList"]
            self.i10.delete(0, "end")
            self.i10.insert(0, dExcelInfo["desc"])

    # CLASS FUNCTION : Save edit excel info
    def SaveExcelInfo(self):
        if self.fileName in self.dListFlag:
            orgExcelInfo = self.dListFlag[self.fileName]
        else:
            orgExcelInfo = self.dExcelList[self.fileName]
            
        textTag = self.i6.get()
        textDesc = self.i10.get()
        
        if orgExcelInfo["tag"] == textTag and orgExcelInfo["desc"] == textDesc:
            if self.fileName in self.dListFlag:
                del self.dListFlag[self.fileName]
            
            self.dExcelList[self.fileName] = orgExcelInfo
        else:   
            if self.fileName not in self.dListFlag:
                self.dListFlag[self.fileName] = copy.deepcopy(orgExcelInfo)
                self.UpdateChangeFlag

            dExcelInfo = self.dExcelList[self.fileName]
            dExcelInfo["tag"] = textTag
            dExcelInfo["desc"] = textDesc
            self.dExcelList[self.fileName] = dExcelInfo
        
        self.UpdateChangeFlag()
    
    # CLASS FUNCTION : Update change flag
    def UpdateChangeFlag(self):
        for idx in self.listFile.get_children():
            fileName = self.listFile.item(idx)["values"][0]
            if fileName in self.dListFlag:
                self.listFile.item(idx, values=(fileName, "CHANGED"))
            else:
                self.listFile.item(idx, values=(fileName, ""))
        
    # CLASS FUNCTION : Click save button
    def Save(self):
        self.SaveExcelInfo()
        self.state = EditPopupState.SAVE
        self.top.destroy()        
    
    # CLASS FUNCTION : Click exit button
    def Exit(self):
        self.state = EditPopupState.EXIT
        self.top.destroy()

# --------------------------------------------------------------------------------------------------------------
class cConfigPopup(object):
    def __init__(self, master):
        self.state = ""
        self.mainUI = master
        # Top-level window 설정
        top = self.top = Toplevel(master.main)
        top.geometry("600x400")
        top.title("Settings")
        #top.wm_overrideredirect(True)  # 기본 창 컨트롤 제거

        if hasattr(sys, "_MEIPASS"):
            icon_path = os.path.join(sys._MEIPASS, "icon/excelexporter.ico")
        else:
            icon_path = "icon/excelexporter.ico"
        if os.path.isfile(icon_path):
            top.iconbitmap(icon_path)

        top.config(bg="#E0E0E0")

        # 창을 최상단으로 유지 및 포커스 설정
        top.lift()
        top.focus_force()
        top.grab_set()

        # 레이아웃 구성
        self.f1 = Frame(top, bd=0, bg="#E0E0E0")
        self.f2 = Frame(top, bd=0, bg="#E0E0E0")
        self.f3 = Frame(top, bd=0, bg="#E0E0E0")

        self.f1.grid(row=0, column=0, padx=8, pady=8, sticky=NSEW)
        self.f2.grid(row=1, column=0, padx=8, pady=8, sticky=NSEW)
        self.f3.grid(row=2, column=0, padx=4, pady=4, sticky=NSEW)

        # 상단 그룹 제목
        self.upperGroupLabel = Label(self.f1, text="Export Config", font=("Arial", 12, "bold"), bg="#E0E0E0",
                                     anchor=CENTER)
        self.upperGroupLabel.grid(row=0, column=0, columnspan=2, padx=4, pady=4, sticky=NSEW)

        # 상단 파일 경로 입력 필드 및 레이블
        self.lNamespace = Label(self.f1, text="Namespace :", width=18, anchor=W, font=("Arial", 10, "bold"), bg="#E0E0E0")
        self.lTablePath = Label(self.f1, text="TablePath :", width=18, anchor=W, font=("Arial", 10, "bold"), bg="#E0E0E0")
        self.lScriptPath = Label(self.f1, text="ScriptPath :", width=18, anchor=W, font=("Arial", 10, "bold"), bg="#E0E0E0")

        self.eNamespace = Entry(self.f1)
        self.eTablePath = Entry(self.f1)
        self.eScriptPath = Entry(self.f1)

        self.eNamespace.insert(0, self.mainUI.strNamespace)
        self.eTablePath.insert(0, self.mainUI.strTablePath)
        self.eScriptPath.insert(0, self.mainUI.strScriptPath)

        self.lNamespace.grid(row=1, column=0, padx=4, pady=4, sticky=NSEW)
        self.lTablePath.grid(row=2, column=0, padx=4, pady=4, sticky=NSEW)
        self.lScriptPath.grid(row=3, column=0, padx=4, pady=4, sticky=NSEW)

        self.eNamespace.grid(row=1, column=1, padx=4, pady=4, sticky=NSEW)
        self.eTablePath.grid(row=2, column=1, padx=4, pady=4, sticky=NSEW)
        self.eScriptPath.grid(row=3, column=1, padx=4, pady=4, sticky=NSEW)

        # 하단 그룹 제목
        self.lowerGroupLabel = Label(self.f2, text="Local Config", font=("Arial", 12, "bold"), bg="#E0E0E0",
                                     anchor=CENTER)
        self.lowerGroupLabel.grid(row=0, column=0, columnspan=2, padx=4, pady=4, sticky=NSEW)

        # 하단 파일 경로 입력 필드 및 레이블
        self.lProjectPath = Label(self.f2, text="ProjectFullPath :", width=18, anchor=W, font=("Arial", 10, "bold"), bg="#E0E0E0")
        #self.lExecuteUnityPath = Label(self.f2, text="ExecuteUnityFullPath :", width=18, anchor=W, font=("Arial", 10, "bold"), bg="#E0E0E0")

        self.eProjectPath = Entry(self.f2)
        #self.eExecuteUnityPath = Entry(self.f2)

        self.eProjectPath.insert(0, self.mainUI.strProjectFullPath)
        #self.eExecuteUnityPath.insert(0, self.mainUI.strExecuteUnityFullPath)

        self.lProjectPath.grid(row=1, column=0, padx=4, pady=4, sticky=NSEW)
        #self.lExecuteUnityPath.grid(row=2, column=0, padx=4, pady=4, sticky=NSEW)

        self.eProjectPath.grid(row=1, column=1, padx=4, pady=4, sticky=NSEW)
        #self.eExecuteUnityPath.grid(row=2, column=1, padx=4, pady=4, sticky=NSEW)

        # 저장 및 취소 버튼
        self.btnSave = Button(self.f3, text="SAVE", width=20, height=2, command=self.Save)
        self.btnExit = Button(self.f3, text="CANCEL", width=20, height=2, command=self.Exit)

        self.btnSave.grid(row=0, column=0, padx=4, pady=4, sticky=NSEW)
        self.btnExit.grid(row=0, column=1, padx=4, pady=4, sticky=NSEW)

        self.f1.rowconfigure((0,1,2,3), weight=1)
        self.f1.columnconfigure( 1, weight=1)

        self.f2.rowconfigure((0,1,2), weight=1)
        self.f2.columnconfigure( 1, weight=1)

        # 그리드 크기 설정
        self.f3.rowconfigure(0, weight=1)
        self.f3.columnconfigure((0, 1), weight=1)

        top.rowconfigure((0, 1, 2), weight=1)
        top.columnconfigure(0, weight=1)
        Center(top)

    def Save(self):

        self.mainUI.strNamespace = self.eNamespace.get()
        self.mainUI.strTablePath = self.eTablePath.get()
        self.mainUI.strScriptPath= self.eScriptPath.get()

        self.mainUI.strProjectFullPath = self.eProjectPath.get()
        #self.mainUI.strExecuteUnityFullPath = self.eExecuteUnityPath.get()

        self.mainUI.SaveConfigData()
        self.top.destroy()

    def Exit(self):
        """ 창 닫기 """
        self.top.destroy()
# --------------------------------------------------------------------------------------------------------------
# CLASS : UI Parameter
class cUI:
    def __init__(self):
        # Parameters
        self.dShowList = dict()
        self.dExcelList = dict()
        self.dSheetList = dict()
        self.dMergeList = dict()
        self.dMergedFileList = dict()
        self.dCheckedFileList = list()
        self.dPreset = dict()
        self.dButtonList = list()
        self.state = MainUIState.DISABLED

        #exportConfig
        self.strNamespace = ""
        self.strTablePath = "Assets/Application/Bundles/Tables/Generated/"
        self.strScriptPath = "Assets/Application/Scripts/Tables/Generated/"

        #localConfig
        self.strProjectFullPath = ""
        self.strExecuteUnityFullPath = ""

        self.cPath = os.getcwd()
        self.excelPath = os.getcwd() + _SEP + "excel" + _SEP
        self.configPath = os.getcwd() + _SEP + "config" + _SEP

        self.exportConfigFile = os.getcwd() + _SEP + "config" + _SEP + "export_config.json"
        self.localConfigFile = os.getcwd()  + _SEP + "config" + _SEP + "local_config.json"

        self.configFile = os.getcwd() + _SEP + "config" + _SEP + "_excel_list.json"
        self.presetFile = os.getcwd() + _SEP + "config" + _SEP + "_preset.json"


        os.makedirs(os.path.dirname(self.excelPath), exist_ok=True)
        os.makedirs(os.path.dirname(self.configPath), exist_ok = True)

        # UI main
        self.main = Tk()

        self.main.title("TABLE SCRIPT")
        self.main.geometry("1280x720")

        if hasattr(sys, "_MEIPASS"):
            icon_path = os.path.join(sys._MEIPASS, "icon/excelexporter.ico")
        else:
            icon_path = "icon/excelexporter.ico"

        if os.path.isfile(icon_path):
            self.main.iconbitmap(icon_path)

        self.main.bind("<FocusIn>", self.OnFocusIn)

        # UI Frame
        self.fTButton = Frame(self.main, bd=0)
        self.fFilter = Frame(self.main, bd=0)
        self.fExcelList = Frame(self.main, bd=0)
        self.fBButton = Frame(self.main, bd=0)
        self.fStatus = Frame(self.main, bd=0)
        
        self.fTButton.grid(row=0, column=0, padx=4, pady=4, sticky=NSEW)
        self.fFilter.grid(row=1, column=0, padx=4, pady=4, sticky=NSEW)
        self.fExcelList.grid(row=2, column=0, padx=4, pady=4, sticky=NSEW)
        self.fBButton.grid(row=3, column=0, padx=4, pady=4, sticky=NSEW)
        self.fStatus.grid(row=4, column=0, padx=4, pady=4, sticky=NSEW)
        
        # Top buttons
        self.btnCheckAll = Button(self.fTButton, width=60, height=1, text='Check All', font='sans 12 bold', command=self.CheckAll)
        self.btnUncheckAll = Button(self.fTButton, width=60, height=1, text='Uncheck All', font='sans 12 bold', command=self.UncheckAll)
        self.btnCheckType = Button(self.fTButton, width=60, height=1, text='Check Type', font='sans 12 bold', command=self.CheckType)

        self.autoSaveToProjectVar = IntVar()
        self.autoSaveToProjectVar.set(1)
        #self.chkAutoSaveToProject = Checkbutton(self.fTButton, text="Auto Save Exported Files to Project", variable=self.autoSaveToProjectVar, command=self.SaveConfigData, anchor=W)
        #self.btnCheckLocaltext = Button(self.fTButton, width=60, height=2, text='Check Localtext', font='sans 12 bold', command=self.CheckLocaltext)
        self.btnRun = Button(self.fTButton, width=90, height=1, text='EXPORT', font='sans 12 bold', bg="#fff2cc", command=lambda :ExportJsonMain(self))
        
        self.btnCheckAll.grid(row=0, column=0, padx=4, pady=4, sticky=NSEW)
        self.btnUncheckAll.grid(row=1, column=0, padx=4, pady=4, sticky=NSEW)
        self.btnCheckType.grid(row=2, column=0, padx=4, pady=4, sticky=NSEW)
       # self.chkAutoSaveToProject.grid(row=0, column=1, padx=4, pady=4, sticky=NSEW)
        #self.btnCheckLocaltext.grid(row=1, column=1, padx=4, pady=4, sticky=NSEW)
        self.btnRun.grid(row=0, column=1, rowspan=3, padx=4, pady=4, sticky=NSEW)
        
        self.dButtonList.append(self.btnCheckAll)
        self.dButtonList.append(self.btnUncheckAll)
        self.dButtonList.append(self.btnCheckType)
        #self.dButtonList.append(self.btnCheckLocaltext)
        self.dButtonList.append(self.btnRun)
        
        self.fTButton.grid_columnconfigure((0,1,2), weight=1)
        self.fTButton.grid_rowconfigure((0,1), weight=1)

        # UI Filters
        self.filterShowAllVar = IntVar()
        self.filterShowAll = Checkbutton(self.fFilter, text="Show All Excels", variable=self.filterShowAllVar, command=self.RefreshListExcel, anchor=W)
        self.filterShowAll.grid(row=0, column=0, padx=4, pady=4, sticky=NSEW)

        #self.branchLabel = Label(self.fFilter, text="Current Branch : ", anchor=E)
        #self.branchName = Label(self.fFilter, text= current_branch, relief=SUNKEN, anchor=W)
        #self.branchLabel.grid(row=0, column=1, padx=4, pady=4, sticky=NSEW)
        #self.branchName.grid(row=0, column=2, padx=4, pady=4, sticky=NSEW)

       # self.filter = Entry(self.fFilter, width=100)
       # self.btnFilter = Button(self.fFilter, width=30, height=1, text="Filter", font="sans 10 bold", command=self.ApplyFilter)
       # self.btnClear = Button(self.fFilter, width=30, height=1, text="Filter Clear", font="sans 10 bold", command=self.FilterClear)
       # self.btnPresetSave = Button(self.fFilter, width=30, height=1, text="Preset Save", font="sans 10 bold", command=self.PopupSavePreset)
       # self.btnPresetLoad = Button(self.fFilter, width=30, height=1, text="Preset Load", font="sans 10 bold", command=self.PopupLoadPreset)
        
       # self.filter.grid(row=0, column=0, padx=4, pady=4, sticky=NSEW)
       # self.btnFilter.grid(row=0, column=1, padx=4, pady=4, sticky=NSEW)
       # self.btnClear.grid(row=0, column=2, padx=4, pady=4, sticky=NSEW)
       # self.btnPresetSave.grid(row=0, column=3, padx=4, pady=4, sticky=NSEW)
       # self.btnPresetLoad.grid(row=0, column=4, padx=4, pady=4, sticky=NSEW)
        
       # self.dButtonList.append(self.btnFilter)
       # self.dButtonList.append(self.btnClear)
       # self.dButtonList.append(self.btnPresetSave)
       # self.dButtonList.append(self.btnPresetLoad)
        
        self.fFilter.grid_columnconfigure((0,1,2,3,4,5,6,7), weight=1)
        self.fFilter.grid_rowconfigure(0, weight=1)

        # Excel file list
        self.scrollExcel = Scrollbar(self.fExcelList)
        self.scrollExcel.pack(side=RIGHT, fill=Y)
        
        style = Style()
        style.theme_use("default")
        style.configure("Treeview", rowheight=25)
        style.configure("Treeview.Heading", font='sans 10 bold', background="#B0B0B0")

        self.listExcel = CheckboxTreeview(self.fExcelList, yscrollcommand=self.scrollExcel.set, selectmode="extended")
        self.listExcel.pack(side=LEFT, fill=BOTH, expand=YES)
        self.scrollExcel.config(command=self.listExcel.yview)
        
        #self.listExcel["columns"] = ("index", "tableType", "tag", "subDir", "fileName", "sheetList", "desc")
        #("index", "tableType", "mergeName", "mergeGroup", "mergeCount", "tag", "subDir", "fileName", "sheetList", "desc")
        self.listExcel["columns"] = ("index", "tableType", "filePath", "sheetList")

        self.listExcel.column("#0", width=50, anchor=CENTER, stretch=NO)
        self.listExcel.column("index", width=50, anchor=CENTER, stretch=NO)
        self.listExcel.column("tableType", width=100, anchor=CENTER, stretch=NO)
        #self.listExcel.column("mergeName", width=100, anchor=CENTER, stretch=NO)
        #self.listExcel.column("mergeGroup", width=100, anchor=CENTER, stretch=NO)
        #self.listExcel.column("mergeCount", width=100, anchor=CENTER, stretch=NO)
        #self.listExcel.column("tag", width=100, anchor=W)
        #self.listExcel.column("subDir", width=150, anchor=W)
        self.listExcel.column("filePath", width=250, anchor=W)
        self.listExcel.column("sheetList", width=300, anchor=W)
        #self.listExcel.column("desc", width=300, anchor=W)
        
        self.listExcel.heading("#0", text="CHK", anchor=CENTER)
        self.listExcel.heading("index", text="NO", anchor=CENTER)
        self.listExcel.heading("tableType", text="TYPE", anchor=CENTER)
        #self.listExcel.heading("mergeName", text="MERGE NAME", anchor=CENTER)
        #self.listExcel.heading("mergeGroup", text="MERGE GROUP", anchor=CENTER)
        #self.listExcel.heading("mergeCount", text="MERGE COUNT", anchor=CENTER)
        #self.listExcel.heading("tag", text="TAG", anchor=CENTER)
        #self.listExcel.heading("subDir", text="SUB DIR", anchor=CENTER)
        self.listExcel.heading("filePath", text="FILE PATH", anchor=CENTER)
        self.listExcel.heading("sheetList", text="DATA SHEETS", anchor=CENTER)
        #self.listExcel.heading("desc", text="DESC", anchor=CENTER)
        
        self.listExcel.tag_configure("checked", background="#00FF00")
        self.listExcel.bind("<<TreeviewSelect>>", self.EventSelect)

        # UI Bottom Buttons
        #self.btnRemove = Button(self.fBButton, width=20, height=1, text="Remove excel files", font='sans 10 bold', command=self.RemoveExcelFiles)
        #self.btnAdd = Button(self.fBButton, width=20, height=1, text="Add excel files", font='sans 10 bold', command=self.AddExcelFiles)
        #self.btnUpdate = Button(self.fBButton, width=20, height=1, text="Update sheet list", font='sans 10 bold', command=self.ForceUpdateExcelFiles)
        self.btnSettings = Button(self.fBButton, width=20, height=1, text="Settings", font='sans 10 bold', command=self.PopupSettings)

        #self.btnRemove.pack(side=LEFT, fill=Y, padx=4, pady=4)
        #self.btnAdd.pack(side=LEFT, fill=Y, padx=4, pady=4)
        #self.btnUpdate.pack(side=LEFT, fill=Y, padx=4, pady=4)

        self.btnSettings.pack(side=RIGHT, fill=Y, padx=4, pady=4)

        #self.dButtonList.append(self.btnRemove)
        #self.dButtonList.append(self.btnAdd)
        #self.dButtonList.append(self.btnUpdate)
        self.dButtonList.append(self.btnSettings)

        # UI Status Bar
        self.status = Label(self.fStatus, text="Table Script", bd=1, relief=SUNKEN, anchor=E)

        self.status.grid(row=0, column=0, columnspan=5, padx=4, pady=4, sticky=NSEW)
        
        self.fStatus.grid_columnconfigure(0, weight=1)
        self.fStatus.grid_rowconfigure(0, weight=1)

        self.main.grid_columnconfigure((0,1,2,3,4), weight=1)
        self.main.grid_rowconfigure(2, weight=1)
        
        self.state = MainUIState.NORMAL
        
        Center(self.main)

    # UI FUNCTION : Click event
    def EventSelect(self, event):
        if self.state == MainUIState.DISABLED:
            self.RestoreChecked()
            self.main.update()
        else:
            sel = self.listExcel.focus()
            if len(sel) > 0:
                fileName = str(os.path.basename(self.listExcel.item(sel)["values"][2])) #merge 컬럼을 임시로 제거해서 인덱스 번호 수정 7 => 4
                
                if "checked" in self.listExcel.item(sel)["tags"]:
                    if fileName in self.dCheckedFileList:
                        self.listExcel.change_state(sel, "unchecked")
                else:
                    if fileName not in self.dCheckedFileList:
                        self.listExcel.change_state(sel, "checked")
            
            self.UpdateChecked()

    # UI FUNCTION : Update Statusbar text = count checked files
    def UpdateStatusBarSelectedFiles(self):
        sizeList = len(self.listExcel.get_checked())
        
        if sizeList == 0:
            self.ChangeStatus("no file selected ( 0 / " + str(len(self.dExcelList)) + " )")
        elif sizeList == 1:
            self.ChangeStatus(str(sizeList) + " file selected ( " + str(sizeList) + " / " + str(len(self.dExcelList)) + " )")
        else:
            self.ChangeStatus(str(sizeList) + " files selected ( " + str(sizeList) + " / " + str(len(self.dExcelList)) + " )")
        
        self.main.update()

    # UI Function : Update status bar text
    def ChangeStatus(self, v: str):
        self.status["text"] = v
        self.main.update()

    # UI FUNCTION : Update all excel files sheetlist & info
    def UpdateSheetList(self):
        self.dSheetList = dict()
        
        for fileName in self.dExcelList:
            sList = self.dExcelList[fileName]["sheetList"]
            for sheetName in sList:
                self.dSheetList[sheetName] = fileName
    
    # UI FUNCTION : Update all excel files merge info
    def UpdateMergeList(self):
        self.dMergeList = dict()
        self.dMergedFileList = dict()
        
        for fileName in self.dExcelList:
            if "mergeName" in self.dExcelList[fileName]:
                mergeName = self.dExcelList[fileName]["mergeName"]
                if len(mergeName) > 0:
                    self.dMergedFileList[fileName] = mergeName
                    mList = list()
                    if mergeName in self.dMergeList:
                        mList = self.dMergeList[mergeName]

                    mList.append(fileName)
                    self.dMergeList[mergeName] = mList

    # UI FUNCTION : Load config data from json file
    def LoadConfigData(self):
        loadFail = False
        if os.path.isfile(self.exportConfigFile) and os.access(self.exportConfigFile, os.R_OK):
            with open(self.exportConfigFile, "r") as file:
                loadData = json.load(file)
                if "Namespace" in loadData:
                    self.strNamespace = loadData["Namespace"]
                if "TablePath" in loadData:
                    self.strTablePath = loadData["TablePath"]
                if "ScriptPath" in loadData:
                    self.strScriptPath = loadData["ScriptPath"]
        else:
            loadFail = True

        if os.path.isfile(self.localConfigFile) and os.access(self.localConfigFile, os.R_OK):
            with open(self.localConfigFile, "r") as file:
                loadData = json.load(file)
                if "ProjectFullPath" in loadData:
                    self.strProjectFullPath = loadData["ProjectFullPath"]
                #if "ExecuteUnityFullPath" in loadData:
                #    self.strExecuteUnityFullPath= loadData["ExecuteUnityFullPath"]
                #if "AutoSaveExportToProject" in loadData:
                #    self.autoSaveToProjectVar.set(loadData["AutoSaveExportToProject"])

        else:
            loadFail = True
            git_root = find_git_root()
            if git_root:
                for root, dirs, files in os.walk(git_root):
                    if is_unity_project(root):
                        self.strProjectFullPath = root
                        break

        if loadFail :
            self.SaveConfigData()


    # UI FUNCTION : Save config data to json file
    def SaveConfigData(self):
        #saveData["ExcelList"] = self.dExcelList
        exportConfigData = dict()
        localConfigData = dict()
        #exportConfigFile
        exportConfigData["Namespace"] = self.strNamespace
        exportConfigData["TablePath"] = self.strTablePath
        exportConfigData["ScriptPath"] = self.strScriptPath

        #localConfigFile
        localConfigData["ProjectFullPath"] = self.strProjectFullPath
        #localConfigData["ExecuteUnityFullPath"] = self.strExecuteUnityFullPath
        #localConfigData["AutoSaveExportToProject"] = self.autoSaveToProjectVar.get()

        os.makedirs(os.path.dirname(self.exportConfigFile), exist_ok = True)
        with open(self.exportConfigFile, "w", encoding="UTF-8") as file:
            json.dump(exportConfigData, file)

        os.makedirs(os.path.dirname(self.localConfigFile), exist_ok = True)
        with open(self.localConfigFile, "w", encoding="UTF-8") as file:
            json.dump(localConfigData, file)


    # UI FUNCTION: Load preset data file
    def LoadPresetData(self):
        if os.path.isfile(self.presetFile) and os.access(self.configFile, os.R_OK):
            with open(self.presetFile, "r") as file:
                self.dPreset = json.load(file)
        else:
            self.SavePresetData()

    # UI FUNCTION : Save preset data file
    def SavePresetData(self):
        os.makedirs(os.path.dirname(self.presetFile), exist_ok = True)
        with open(self.presetFile, "w", encoding="UTF-8") as file:
            json.dump(self.dPreset, file)

    # UI Function : Update chekced file list
    def UpdateChecked(self):
        self.dCheckedFileList = list()
            
        for idx in self.listExcel.get_checked():
            fileName = self.GetExcelInfoByIndex(idx)["fileName"]
            if fileName in self.dExcelList:
                self.dCheckedFileList.append(fileName)
                
        self.UpdateStatusBarSelectedFiles()

    # UI Function : Restore Checkbox file list by checked file list
    def RestoreChecked(self):
        cList = list()
        for fileName in self.dCheckedFileList:
            if fileName in self.dExcelList:
                cList.append(fileName)
        
        self.dCheckedFileList = cList
        cList = self.listExcel.get_checked()
        
        for fileName in self.dExcelList:
            idx = self.dExcelList[fileName]["index"]            
            if fileName in self.dCheckedFileList:
                if str(idx) not in cList:
                    if self.listExcel.exists(idx) :
                        self.listExcel.change_state(idx, "checked")
            else:
                if str(idx) in cList:
                    self.listExcel.change_state(idx, "unchecked")
                    
        self.UpdateStatusBarSelectedFiles()

    # UI FUNCTION : Check All files
    def CheckAll(self):
        self.ChangeCheck("ALL", True)

    # UI FUNCTION : Uncheck All files
    def UncheckAll(self):
        self.ChangeCheck("ALL", False)

    # JSON FUNCTION : Check types
    def CheckType(self):
        self.ChangeCheck("TYPE", True)

    # UI FUNCTION : Check localtext
    def CheckLocaltext(self):
        self.ChangeCheck("LOCALTEXT", True)

    # UI FUNCTION : Change check/uncheck state by condition
    def ChangeCheck(self, t: str, c: bool):
        for idx in self.listExcel.get_children():
            excelInfo = self.GetExcelInfoByIndex(idx)
            if t == excelInfo["type"] or t == "ALL":
                if c:
                    self.listExcel.change_state(idx, "checked")
                else:
                    self.listExcel.change_state(idx, "unchecked")
        
        self.UpdateChecked()
        # self.UpdateStatusBarSelectedFiles()

    # UI FUNCTION : Get Excel Info by table idx
    def GetExcelInfoByIndex(self, idx):
        filename = str(os.path.basename(self.listExcel.item(idx)["values"][2]))
        return self.dExcelList[filename] #merge 컬럼을 임시로 제거해서 인덱스 번호 수정 7 => 4

    # UI FUNCTION : Sort excel file list (filename / sub dir)
    def SortTableList(self):
        d1 = dict()

        for fileName in self.dExcelList:
            d1[fileName] = self.dExcelList[fileName]["subDir"]
        
        d2 = dict(sorted(d1.items()))
        d1 = dict(sorted(d2.items(), key=lambda item: item[1]))
        
        listType = list()
        listLocaltext = list()
        listMerge = list()
        listData = list()
        
        for fileName in d1:
            match(self.dExcelList[fileName]["type"]):
                case "TYPE": listType.append(fileName)
                case "LOCALTEXT": listLocaltext.append(fileName)
                case "MERGE": listMerge.append(fileName)
                case _: listData.append(fileName)
        
        listAll = listType + listLocaltext + listMerge + listData
        
        nExcelList = dict()
        
        for fileName in listAll:
            oExcelInfo = self.dExcelList[fileName]
            nExcelInfo = dict()
            nExcelInfo["index"] = len(nExcelList) + 1
            nExcelInfo["type"] = oExcelInfo["type"]
            # nExcelInfo["mergeName"] = oExcelInfo["mergeName"]
            # nExcelInfo["mergeGroup"] = oExcelInfo["mergeGroup"]
            # nExcelInfo["mergeCount"] = oExcelInfo["mergeCount"]
            nExcelInfo["tag"] = oExcelInfo["tag"]
            nExcelInfo["fullPath"] = oExcelInfo["fullPath"]
            nExcelInfo["subDir"] = oExcelInfo["subDir"]
            nExcelInfo["fileName"] = oExcelInfo["fileName"]
            nExcelInfo["sheetList"] = oExcelInfo["sheetList"]
            nExcelInfo["desc"] = oExcelInfo["desc"]
            nExcelList[fileName] = nExcelInfo
        
        self.dExcelList = nExcelList

    def RefreshGitChangeListExcel(self):
        current_path = os.getcwd().replace("\\","/")
        folder_path = 'excel/'  # 변경 여부를 확인할 파일 경로
        all_files = get_all_files_in_folder(folder_path)
        for file in all_files:
            print(file)

        self.dShowList.clear()
        git_root = find_git_root()
        if git_root:
            git_root = git_root.replace("\\","/")  + "/"
            git_path = current_path.removeprefix(git_root)
            repo = git.Repo(search_parent_directories=True)
            changed_files = [item.a_path for item in repo.index.diff(None)]
            added_files = repo.untracked_files  # Unstaged changes
            staged_files = [item.a_path for item in repo.index.diff('HEAD')]  # Staged changes

            for file in all_files:
                file_path = git_path + "/"+ file
                if self.CheckFileChanged(file_path, changed_files, staged_files, added_files):
                    fileName = os.path.basename(file_path)
                    if fileName in self.dExcelList:
                        dExcelInfo = self.dExcelList[fileName]
                        remainPath = dExcelInfo["fullPath"].removeprefix(git_root)
                        if remainPath == file_path:
                            self.dShowList[fileName] = dExcelInfo

    def CheckFileChanged(self, file_path, changed_files, staged_files, added_files):
            # 'unstaged' 또는 'staged' 상태에 있는 파일을 검색
            return file_path in changed_files or file_path in staged_files or file_path in added_files

    # UI FUNCTION : Refresh excel file list
    def RefreshListExcel(self):
        #self.SortTableList()

        # Clear List
        for item in self.listExcel.get_children():
            self.listExcel.delete(item)

        if self.filterShowAllVar.get() == 1:
            self.dShowList = self.dExcelList.copy()
        else :
            self.RefreshGitChangeListExcel()
        #if len(self.filter.get()) == 0:
        #    self.dShowList = self.dExcelList

        # Update List
        for fileName in self.dShowList:
            dExcelInfo = self.dExcelList[fileName]
            self.listExcel.insert(parent="", index="end", iid=dExcelInfo["index"], values=(dExcelInfo["index"], dExcelInfo["type"], dExcelInfo["subDir"]+dExcelInfo["fileName"], dExcelInfo["sheetList"]))

            #self.listExcel.insert(parent="", index="end", iid=dExcelInfo["index"], values=(dExcelInfo["index"], dExcelInfo["type"], dExcelInfo["tag"], dExcelInfo["subDir"], dExcelInfo["fileName"], dExcelInfo["sheetList"], dExcelInfo["desc"]))
            #self.listExcel.insert(parent="", index="end", iid=dExcelInfo["index"], values=(dExcelInfo["index"], dExcelInfo["type"], dExcelInfo["mergeName"], dExcelInfo["mergeGroup"], dExcelInfo["mergeCount"], dExcelInfo["tag"], dExcelInfo["subDir"], dExcelInfo["fileName"], dExcelInfo["sheetList"], dExcelInfo["desc"]))
        
        self.RestoreChecked()
        self.UpdateChecked()
        self.main.update()    

    # UI FUNCTION : Apply filter
    def ApplyFilter(self):
        strFilter = str(self.filter.get())
        self.dShowList = dict()
        if strFilter == "":
            self.dShowList = self.dExcelList
        else:
            for fileName in self.dExcelList:
                dExcelInfo = self.dExcelList[fileName]
                
                for sheetName in dExcelInfo["sheetList"]:
                    if strFilter in sheetName:
                        self.dShowList[fileName] = dExcelInfo
                        break
                
                if strFilter.lower() in str(dExcelInfo["tag"]).lower() or strFilter.lower() in str(dExcelInfo["fileName"]).lower():
                    self.dShowList[fileName] = dExcelInfo
                elif fileName in self.dCheckedFileList:
                    self.dShowList[fileName] = dExcelInfo
        
        self.RefreshListExcel()

    # UI FUNCTION : Clear filter
    def FilterClear(self):
        self.filter.delete(0, END)
        self.RefreshListExcel()

    # UI FUNCTION : Edit Export Settings
    def PopupSettings(self):
        self.ButtonDisable()
        p = cConfigPopup(self)
        self.main.wait_window(p.top)

        #if len(self.dCheckedFileList) > 0:
        #    p = cEditPopup(self)
        #    self.main.wait_window(p.top)
                
        #    match(p.state):
        #        case EditPopupState.SAVE:
        #            self.dExcelList = p.dExcelList
        #            self.SaveConfigData()
        #            self.RefreshListExcel()
        #else:
            #messagebox.showwarning("Edit excel info popup", "No checked files")

        self.ButtonEnable()

    # UI FUNCTION : Add excel files
    def AddExcelFiles(self):
        self.ButtonDisable()
        addFileList = filedialog.askopenfilenames(initialdir="./", title="Select Excel Files", filetypes=[("Excel files", ".xls .xlsx .xlsm")])
        
        if len(addFileList) > 0:
            pBar = cProgressPopup(self)
            maxRate = len(addFileList)
            pBar.SetInit("Adding selected excel files", "", "0 / " + str(maxRate), 0)
            count = 0
            
            for fullPath in addFileList:
                fileName = str(os.path.basename(fullPath))
                fileDir = str(os.path.dirname(fullPath)).replace("\\", "/")
                str1 = str(self.cPath).replace("\\", "/")
                if fileDir != str1:
                    fileDir = fileDir.replace(str1 + "/", "")
                else:
                    fileDir = ""
                currentPath = "./" + fileDir + "/" + fileName
                mName = ""
                mGroup = ""
                mCount = 0

                pBar.SetText(fileName, str(count) + " / " + str(maxRate))

                wb = load_workbook(filename = fullPath, read_only = True, data_only = True)
                sList = list()
                dataType = "NULL"
                for ws in wb.worksheets:
                    if ws.max_row > 2:
                        match(ws.cell(2,1).value):
                            case "#type":
                                dataType = "TYPE"
                                sList.append(ws.title)
                            case "#localtext":
                                dataType = "LOCALTEXT"
                                sList.append(ws.title)
                            case "#merge":
                                listMergeData = ParseToList(ws.cell(3,1).value, ",")
                                if len(listMergeData) == 3:
                                    mName = listMergeData[0]
                                    mGroup = listMergeData[1]
                                    mCount = listMergeData[2]
                                dataType = "MERGE"
                                sList.append(ws.title)
                            case "#data":
                                if dataType == "NULL": dataType = "DATA"
                                sList.append(ws.title)
                
                if len(sList) > 0:
                    if fileName in self.dExcelList:
                        self.dExcelList[fileName]["sheetList"] = sList
                    else:
                        dExcelInfo = dict()                
                        dExcelInfo["index"] = len(self.dExcelList) + 1
                        dExcelInfo["type"] = dataType
                        #dExcelInfo["mergeName"] = mName
                        #dExcelInfo["mergeGroup"] = mGroup
                        #dExcelInfo["mergeCount"] = mCount
                        dExcelInfo["tag"] = ""
                        dExcelInfo["fullPath"] = currentPath
                        dExcelInfo["subDir"] = fileDir
                        dExcelInfo["fileName"] = fileName
                        dExcelInfo["sheetList"] = sList
                        dExcelInfo["desc"] = ""
                        
                        self.dExcelList[fileName] = dExcelInfo
                
                wb.close()
                count += 1
                pBar.SetRate(count / maxRate * 100)
            
            pBar.SetText("Refresh excel list", str(count) + " / " + str(maxRate))
            
            self.UpdateSheetList()
            self.UpdateMergeList()
            self.SaveConfigData()
            self.RefreshListExcel()
            
            pBar.SetRate(100)
            pBar.Exit()
        
        self.ButtonEnable()

    def ForceUpdateExcelFiles(self):
        if self.state == MainUIState.DISABLED:
            return
        print("ForceUpdateExcelFiles")
        self.ButtonDisable()

        current_branch = get_current_branch()
        if not current_branch :
            current_branch = "연결된 깃 브랜치가 없습니다."
            self.filterShowAllVar.set(1)

        self.main.title("TABLE SCRIPT - Current Git Branch : " + current_branch)

        self.dExcelList.clear()
        current_path = os.getcwd()
        folder_path = 'excel/'  # 변경 여부를 확인할 파일 경로
        all_files = get_all_files_in_folder(folder_path)
        for file in all_files:
            _, fileExt = os.path.splitext(file)
            if fileExt == '.xls' or fileExt == '.xlsx' or fileExt == ".xlsm":
                fullPath = (current_path + "/" + file).replace("\\","/")
                fileName = str(os.path.basename(file))
                fileDir = (str(os.path.dirname(file)) + "/").removeprefix("excel/")
                #fileName = str(os.path.basename(fullPath))
                #fileDir = str(os.path.dirname(fullPath)).replace("\\", "/")
                #str1 = str(self.cPath).replace("\\", "/")
                #if fileDir != str1:
                #    fileDir = fileDir.replace(str1 + "/", "")
                #else:
                #    fileDir = ""
                #currentPath = "./" + fileDir + "/" + fileName
                mName = ""
                mGroup = ""
                mCount = 0

                if "~$" in fileName :
                    continue

                if fileName in self.dExcelList:
                   continue

                wb = load_workbook(filename=fullPath, read_only=True, data_only=True)
                sList = list()
                dataType = "NULL"
                for ws in wb.worksheets:
                    if ws.max_row > 2:
                        match (ws.cell(2, 1).value):
                            case "#type":
                                dataType = "TYPE"
                                sList.append(ws.title)
                            case "#localtext":
                                dataType = "LOCALTEXT"
                                sList.append(ws.title)
                            case "#merge":
                                listMergeData = ParseToList(ws.cell(3, 1).value, ",")
                                if len(listMergeData) == 3:
                                    mName = listMergeData[0]
                                    mGroup = listMergeData[1]
                                    mCount = listMergeData[2]
                                dataType = "MERGE"
                                sList.append(ws.title)
                            case "#data":
                                if dataType == "NULL": dataType = "DATA"
                                sList.append(ws.title)

                if len(sList) > 0:
                    if fileName in self.dExcelList:
                        self.dExcelList[fileName]["sheetList"] = sList
                    else:
                        dExcelInfo = dict()
                        dExcelInfo["index"] = len(self.dExcelList) + 1
                        dExcelInfo["type"] = dataType
                        # dExcelInfo["mergeName"] = mName
                        # dExcelInfo["mergeGroup"] = mGroup
                        # dExcelInfo["mergeCount"] = mCount
                        dExcelInfo["tag"] = ""
                        dExcelInfo["fullPath"] = fullPath
                        dExcelInfo["subDir"] = fileDir
                        dExcelInfo["fileName"] = fileName
                        dExcelInfo["sheetList"] = sList
                        dExcelInfo["desc"] = ""

                        self.dExcelList[fileName] = dExcelInfo
                wb.close()
        self.ButtonEnable()
        self.UpdateSheetList()
   #     self.UpdateMergeList()
        self.RefreshListExcel()
    # UI FUNCTION : Update excel files
    def UpdateExcelFiles(self):
        self.ButtonDisable()
        
        if len(self.dExcelList) > 0:
            pBar = cProgressPopup(self)
            maxRate = len(self.dExcelList)
            pBar.SetInit("Updating selected excel files", "", "0 / " + str(maxRate), 0)
            count = 0
            
            nExcelList = dict()
            
            for fileName in self.dExcelList:
                pBar.SetText(fileName, str(count) + " / " + str(maxRate))

                dExcelInfo = self.dExcelList[fileName]
                fullPath = dExcelInfo["fullPath"]
                mName = ""
                mGroup = ""
                mCount = 0                
                
                if os.path.isfile(fullPath):
                    wb = load_workbook(filename = fullPath, read_only = True, data_only = True)
                    sList = list()
                    dataType = "NULL"
                    for ws in wb.worksheets:
                        if ws.max_row > 2:
                            match(ws.cell(2,1).value):
                                case "#type":
                                    dataType = "TYPE"
                                    sList.append(ws.title)
                                case "#localtext":
                                    dataType = "LOCALTEXT"
                                    sList.append(ws.title)
                                case "#merge":
                                    listMergeData = ParseToList(ws.cell(3,1).value, ",")
                                    if len(listMergeData) == 3:
                                        mName = listMergeData[0]
                                        mGroup = listMergeData[1]
                                        mCount = listMergeData[2]
                                    dataType = "MERGE"
                                    sList.append(ws.title)                                    
                                case "#data":
                                    if dataType == "NULL": dataType = "DATA"
                                    sList.append(ws.title)

                    dExcelInfo["type"] = dataType
                    #dExcelInfo["mergeName"] = mName
                    #dExcelInfo["mergeGroup"] = mGroup
                    #dExcelInfo["mergeCount"] = mCount
                    
                    if len(sList) > 0:
                        dExcelInfo["sheetList"] = sList
                    
                    nExcelList[fileName] = dExcelInfo
                    wb.close()
                
                count += 1
                pBar.SetRate(count / maxRate * 100)
            
            pBar.SetText("Refresh excel list", str(count) + " / " + str(maxRate))
            
            self.dExcelList = nExcelList
            self.UpdateSheetList()
            self.UpdateMergeList()
            self.SaveConfigData()
            self.RefreshListExcel()
            
            pBar.SetRate(100)
            pBar.Exit()
        else:
            messagebox.showwarning("Update excel files", "No excel files")
        
        self.ButtonEnable()

    # UI FUNCTION : Remove checked excel files
    def RemoveExcelFiles(self):
        self.ButtonDisable()
        
        if len(self.dCheckedFileList) > 0:
            strTitle = "Remove checked excel files"
            strDesc = "Remove "
            if len(self.dCheckedFileList) == 1:
                strDesc = strDesc + "1 checked excel file?\n"
            else:
                strDesc = strDesc + str(len(self.dCheckedFileList)) + " checked excel file?\n"
            
            cnt = 0
            for fileName in self.dCheckedFileList:
                cnt += 1
                if cnt > 30:
                    strDesc = strDesc + "\n\n    and more files..."
                    break
                strDesc = strDesc + "\n    " + fileName
            
            res = messagebox.askokcancel(strTitle, strDesc)
            
            if res:
                pBar = cProgressPopup(self)
                pBar.SetInit("Removing checked excel files", "Update excel data", "0 / 2", 0)

                nExcelList = dict()
                for fileName in self.dExcelList:
                    if fileName not in self.dCheckedFileList:
                        oExcelInfo = self.dExcelList[fileName]
                        excelInfo = dict()
                        excelInfo["index"] = len(nExcelList) + 1
                        excelInfo["type"] = oExcelInfo["type"]
#                        excelInfo["mergeName"] = oExcelInfo["mergeName"]
#                        excelInfo["mergeGroup"] = oExcelInfo["mergeGroup"]
#                        excelInfo["mergeCount"] = oExcelInfo["mergeCount"]
                        excelInfo["tag"] = oExcelInfo["tag"]
                        excelInfo["fullPath"] = oExcelInfo["fullPath"]
                        excelInfo["subDir"] = oExcelInfo["subDir"]
                        excelInfo["fileName"] = oExcelInfo["fileName"]
                        excelInfo["sheetList"] = oExcelInfo["sheetList"]
                        excelInfo["desc"] = oExcelInfo["desc"]
                        nExcelList[fileName] = excelInfo
                
                self.dExcelList = nExcelList
                
                pBar.SetText("Refresh excel list", "1 / 2")
                pBar.SetRate(50)
                
                self.UpdateSheetList()
                self.UpdateMergeList()
                self.SaveConfigData()
                self.RefreshListExcel()
                
                pBar.SetRate(100)
                pBar.Exit()
        else:
            messagebox.showwarning("Remove excel files", "No checked files")
        
        self.ButtonEnable()

    # UI FUNCTION : Popup save preset by name
    def PopupSavePreset(self):
        self.ButtonDisable()
        strTitle = "Save preset popup"
        strRes = ""
        
        if len(self.dCheckedFileList) > 0:
            p = cPresetSavePopup(self)
            self.main.wait_window(p.top)
            
            match(p.state):
                case PresetSavePopupState.SAVE:
                    if len(p.presetName) == 0:
                        messagebox.showwarning(strTitle, "No preset name")
                    else:
                        if len(self.dCheckedFileList) == 1:
                            strRes = "Save " + str(p.presetName) + " with " + str(len(self.dCheckedFileList)) + " file?"
                        else:
                            strRes = "Save " + str(p.presetName) + " with " + str(len(self.dCheckedFileList)) + " files?"
                        strRes = strRes + "\n--------------------"
                        
                        count = 0
                        for fileName in self.dCheckedFileList:
                            count = count + 1
                            if count > 30:
                                strRes = strRes + "\n\n    and more files..."
                                break
                            else:
                                strRes = strRes + "\n    " + fileName
                    
                        res = messagebox.askokcancel(strTitle, strRes)
                        
                        if res:                    
                            self.dPreset[p.presetName] = self.dCheckedFileList
                            self.SavePresetData()
        else:
            messagebox.showwarning(strTitle, "No checked file")
        
        self.ButtonEnable()

    # UI FUNCTION : Popup load preset by name:
    def PopupLoadPreset(self):
        self.ButtonDisable()
        p = cPresetLoadPopup(self)
        self.main.wait_window(p.top)
        
        match(p.state):
            case PresetLoadPopupState.LOAD:
                if p.presetName in self.dPreset:
                    self.dCheckedFileList = self.dPreset[p.presetName]
                    self.RefreshListExcel()
                else:
                    messagebox.showwarning("Preset load popup", "No preset name")
            case PresetLoadPopupState.NO_NAME:
                messagebox.showwarning("Preset load popup", "No preset name")
            case PresetLoadPopupState.DELETE:
                if p.presetName in self.dPreset:
                    res = messagebox.askokcancel("Delete preset", "Delete " + p.presetName + "?")
                    if res:
                        del self.dPreset[p.presetName]
                        self.SavePresetData()
                    
        self.ButtonEnable()
    
    # UI FUNCTION : All button disable
    def ButtonDisable(self):
        self.ButtonStateChange(DISABLED)
        self.state = MainUIState.DISABLED

    # UI FUNCTION : All button disable
    def ButtonEnable(self):
        self.ButtonStateChange(NORMAL)
        self.state = MainUIState.NORMAL

    # UI FUNCTION : Button state change
    def ButtonStateChange(self, state):
        for btn in self.dButtonList:
            btn["state"] = state

    def OnFocusIn(self, event):
        if event.widget == self.main:
            self.ForceUpdateExcelFiles()


# --------------------------------------------------------------------------------------------------------------
# JSON FUNCTION : UI Call
def GUIMain():    
    u = cUI()
    u.LoadConfigData()
    u.RefreshListExcel()
    u.main.mainloop()

def copy_folder(src_folder, dest_folder):
    # src_folder의 내용을 dest_folder로 복사
    shutil.copytree(src_folder, dest_folder, dirs_exist_ok=True)

def get_all_files_in_folder(folder_path):
    files_list = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            files_list.append(os.path.join(root, file).replace("\\", "/"))
    return files_list

def is_file_changed(file_path):
    repo = git.Repo(search_parent_directories=True)
    # 'unstaged' 또는 'staged' 상태에 있는 파일을 검색
    changed_files = [item.a_path for item in repo.index.diff(None)]
    added_files = repo.untracked_files # Unstaged changes
    staged_files = [item.a_path for item in repo.index.diff('HEAD')]  # Staged changes
    return file_path in changed_files or file_path in staged_files  or file_path in added_files

def get_current_branch():
    # 현재 경로의 git 저장소 불러오기
    try:
        repo = git.Repo(search_parent_directories=True)
    # 현재 브랜치 이름 가져오기
        current_branch = repo.active_branch.name
        return current_branch
    except Exception as e:
        return None

def find_git_root(start_path=None):
    if start_path is None:
        start_path = os.getcwd()  # 현재 작업 디렉토리에서 시작

    current_path = start_path

    while True:
        if os.path.isdir(os.path.join(current_path, '.git')):
            return current_path
        parent_path = os.path.dirname(current_path)
        if parent_path == current_path:  # 루트 디렉토리에 도달했을 경우
            break
        current_path = parent_path
    return None

# pywinauto 라이브러리는 macOS 지원 안함. unity를 확인하거나 띄우지 못함(수동으로 해아함).
if sys.platform != "darwin":
    from pywinauto import Application

def find_unity_editor(project_name):
    if sys.platform == "darwin":
        return None

    app = Application()
    try:
        app.connect(title_re=f".*{project_name}.*Unity.*")
    except Exception as e:
        print(f"유니티 에디터를 찾을 수 없습니다: {e}")
        return None

    unity_windows = app.windows(title_re=f".*{project_name}.*Unity.*")
    return unity_windows

def focus_unity_editor(project_name):
    unity_windows = find_unity_editor(project_name)

    if unity_windows:
        # 첫 번째 유니티 에디터 창을 포커스
        unity_windows[0].set_focus()
    else:
        print("유니티 에디터 창을 찾을 수 없습니다.")


def is_unity_project(folder_path):
    # Unity 프로젝트에 일반적으로 존재하는 폴더 및 파일
    required_folders = ["Assets", "ProjectSettings", "Packages"]
    required_file = os.path.join(folder_path, "ProjectSettings", "ProjectVersion.txt")

    # 필수 폴더가 있는지 확인
    for folder in required_folders:
        if not os.path.isdir(os.path.join(folder_path, folder)):
            return False

    # ProjectVersion.txt 파일이 있는지 확인
    if not os.path.isfile(required_file):
        return False

    # 모든 조건을 충족하면 Unity 프로젝트 폴더임
    return True

def snake_to_pascal(snake_str):
    return ''.join(word[:1].upper() + word[1:] for word in snake_str.split('_'))

if __name__ == '__main__':
    multiprocessing.freeze_support()
    GUIMain()
